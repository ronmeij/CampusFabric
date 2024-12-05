##########
# Name: 02-config-vrf.py
# Vers: 		1.33
# Date: 		241205
# Auth: 		Ronald Meijer
# Pyth: 		version 3
# Dependency:	02-vrf-id-name-isid_RON-SAMPLE_v1.3
# Dependency:	openpyxl
# ChangeLog:	1.32	detects the csv file delimiter
#				1.33	added support for xlsx file type
##########
# Variables:
#	vrf_____id 	vrf id
#	vrf___name 	vrf name
#	vrf___isid 	vrf i-sid
#	vrf__ipvpn 	ipvpn enabled
#	vrf___mvpn 	ip multicast enabled
#	rdstr__dir 	redistribute direct connected interfaces
#	rdstr_stat 	redistribute static routes
#
#	filename	list of the *.csv files in working directory
#	location	specifies working directory
#	fileset		filenames in the list filename
#	outfil		name of the out-file 
#	suffix		'.xlsx'
##########
# Run:      python3 02-config-vrf.py
##########

import csv
import glob
# 'location' specifies working directory
# location = 'c:/test/temp/'
import os.path
import re

# the following may be required
# pip install openpyxl
import openpyxl
from openpyxl import load_workbook
directory = "."

def answer():
	answer = input("Press enter to continue")

print("\nThis script configures VRFs for VOSS:")
print ("\nFile(s) found in local directory:")
fileset = [file for file in glob.glob("02*.xlsx", recursive=False)]


# Print the items with their corresponding indices
for idx, file in enumerate(fileset):
	print(f"    {idx+1}: {file}")

filename =""

while True:
	try:
		selected_index = int(input("Please select an index from the list of files: "))
	except ValueError:
		print("\n\t **** INFORM Please enter a valid integer for the index (a number please...)\n")
		continue
	else:
		# print (selected_index)
		# print (len(fileset)+1)
		if selected_index <= 0:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if selected_index >= len(fileset)+1:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if 1 <= selected_index <= len(fileset)+1:
			selected_file = fileset[selected_index-1]
			# Here you can open or use the file, for example:
			# Open the selected file
			filename = os.path.join(directory, selected_file)
			print("\n\t**** INFORM file found\n")
			print("\nColumn \t Header")
		break
with open (filename) as f:
	workbook = load_workbook(filename)
	# Select the active sheet (or use workbook['SheetName'])
	sheet = workbook.active
	# Start processing from a specific row, e.g., row 3
	header_row_index = 3 
	#for header_row in sheet.iter_rows(min_row=header_row_index, values_only=True):
		#header_row = next(reader)
	header_row = [ cell.value for cell in sheet[header_row_index] if cell.value and str(cell.value).strip() ]
	for index, column_header in enumerate(header_row):
		print(index, "\t", column_header.strip())

	print("\n\t**** INFORM The file " + filename + " was selected\n")
	answer()

	vrf_____id = []
	vrf___name = []
	vrf___isid = []
	vrf__ipvpn = []
	vrf___mvpn = []
	rdstr__dir = []
	rdstr_stat = []

	valid_rows = []

	#for row in reader:
	for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
		if row[0] is not None:
			valid_rows.append(row)

	print('XXXXXXX Nb valid rows=', len(valid_rows))

	for row in valid_rows:
		try:
			vrf_____id.append(row[0] if row[0] is not None else "")
			vrf___name.append(row[1] if row[1] is not None else "")
			vrf___isid.append(row[2] if row[2] is not None else "")
			vrf__ipvpn.append(row[4] if row[4] is not None else "")
			vrf___mvpn.append(row[5] if row[5] is not None else "")
			rdstr__dir.append(row[6] if row[6] is not None else "")
			rdstr_stat.append(row[7] if row[7] is not None else "")
		except IndexError:
			print("\n\n\t**** WARNING excel header not as expected!!!")
			quit()
	vrf_____id = [x.strip(" ") for x in vrf_____id]
	print(vrf_____id)
	vrf___name = [x.strip(" ") for x in vrf___name]
	print(vrf___name)
	vrf___isid = [x.strip(" ") for x in vrf___isid]
	print(vrf___isid)
	vrf__ipvpn = [x.strip(" ") for x in vrf__ipvpn]
	vrf__ipvpn = [x.lower() for x in vrf__ipvpn]
	print(vrf__ipvpn)
	vrf___mvpn = [x.strip(" ") for x in vrf___mvpn]
	vrf___mvpn = [x.lower() for x in vrf___mvpn]
	print(vrf___mvpn)
	rdstr__dir = [x.strip(" ") for x in rdstr__dir]
	rdstr__dir = [x.lower() for x in rdstr__dir]
	print(rdstr__dir)
	rdstr_stat = [x.strip(" ") for x in rdstr_stat]
	rdstr_stat = [x.lower() for x in rdstr_stat]
	print(rdstr_stat)
answer()
lenvrf_____id = len(vrf_____id)
lenvrf___name = len(vrf___name)
lenvrf___isid = len(vrf___isid)
lenvrf__ipvpn = len(vrf__ipvpn)
lenvrf___mvpn = len(vrf___mvpn)
lenrdstr__dir = len(rdstr__dir)
lenrdstr_stat = len(rdstr_stat)
print(lenvrf_____id)
print(lenvrf___name)
print(lenvrf___isid)
print(lenvrf__ipvpn)
print(lenvrf___mvpn)
print(lenrdstr__dir)
print(lenrdstr_stat)
answer()
outfil = "02-out-vrf.cfg"
print("\nThe name of the out-file is " + outfil + " \n")

with open(outfil, 'w') as outfile:
	outfile.write("\n\n")
with open(outfil, 'w') as outfile:
	outfile.write('\n# ***** start of 02-out-vrf'+ '.cfg ***** ')
with open(outfil, 'a') as outfile:
	outfile.write('\n\n')
script = "rwa"
print(script)
with open(outfil, 'a') as outfile:
	outfile.write("rwa\n")
script = "rwa"
print(script)
with open(outfil, 'a') as outfile:
	outfile.write("rwa\n")
script = "enable " + " "
print (script)
with open(outfil, 'a') as outfile:
	outfile.write("enable " + "\n\n")

script = "configure terminal " + " "
print (script)
with open(outfil, 'a') as outfile:
	outfile.write("configure terminal\n\n")

for value in range(0,lenvrf_____id):
	if int(vrf_____id[value]) >= 1:
		script = "ip vrf " + vrf___name[value] + " vrfid " + vrf_____id[value] + " "
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("ip vrf " + vrf___name[value] + " vrfid " + vrf_____id[value] + " \n")
	script = " "
	print(script)
	with open(outfil, 'a') as outfile:
		outfile.write(" \n")

for value in range(0,lenvrf_____id):
	if int(vrf_____id[value]) >= 1:
		script = "router vrf " + vrf___name[value] + " "
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\nrouter vrf " + vrf___name[value] + " \n")
		if "yes" in vrf__ipvpn[value]:
			script = "ipvpn " + " "
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("ipvpn " + " \n")
		script = "i-sid " + vrf___isid[value] + " "
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("i-sid " + vrf___isid[value] + " \n")
		if "yes" in vrf___mvpn[value]:
			script = "mvpn enable " + " "
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("mvpn enable " + " \n")		
		if "yes" in vrf__ipvpn[value]:
			script = "ipvpn enable " + " "
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("ipvpn enable " + " \n")	
		script = "exit " + " "
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("exit " + " \n\n")

for value in range(0,lenvrf_____id):
	if int(vrf_____id[value]) == 0:
		script = 'router isis '
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write('router isis \n')
		if "no" in vrf__ipvpn[value]:
			script = "no spbm 1 ip enable " + " "
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("no spbm 1 ip enable " + " \n")
		if "no" in vrf___mvpn[value]:
			script = "no spbm 1 multicast enable " + " "
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("no spbm 1 multicast enable " + " \n")
		if "yes" in rdstr__dir[value]:
			script = 'redistribute direct'
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('redistribute direct \n')
			script = 'redistribute direct enable '
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('redistribute direct enable \n')
		if "yes" in rdstr_stat[value]:
			script = 'redistribute static'
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('redistribute static \n')
			script = 'redistribute static enable '
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('redistribute static enable \n')
		script = 'exit'
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write('exit \n\n')
		if "yes" in rdstr__dir[value]:
			script = 'isis apply redistribute direct '
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('isis apply redistribute direct \n')
		if "yes" in rdstr_stat[value]:
			script = 'isis apply redistribute static '
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('isis apply redistribute static \n\n')

	if int(vrf_____id[value]) >= 1:
		script = 'router vrf ' + vrf___name[value] + ' '
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write('router vrf ' + vrf___name[value] + ' \n')
		if "yes" in rdstr__dir[value]:
			script = "isis redistribute direct " + " "
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("isis redistribute direct " + " \n")
			script = "isis redistribute direct enable" + " "
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("isis redistribute direct enable" + " \n")
		if "yes" in rdstr_stat[value]:
			script = 'isis redistribute static'
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('isis redistribute static \n')
			script = 'isis redistribute static enable '
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('isis redistribute static enable \n')
		script = 'exit'
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write('exit \n\n')
		if "yes" in rdstr__dir[value]:
			script = 'isis apply redistribute direct vrf ' + vrf___name[value] + ' '
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('isis apply redistribute direct vrf ' + vrf___name[value] + ' \n\n')
		if "yes" in rdstr_stat[value]:
			script = 'isis apply redistribute static vrf ' + vrf___name[value] + ' '
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write('isis apply redistribute static vrf ' + vrf___name[value] + ' \n\n')



script = "\nend " + " "
print(script)
with open(outfil, 'a') as outfile:
	outfile.write("\nend " + "\n\n")
script = "save config " + " "
print (script)
with open(outfil, 'a') as outfile:
	outfile.write("save config " + " \n")
script = "backup configure 02-ip-vrf " + " "
print (script)
with open(outfil, 'a') as outfile:
	outfile.write("backup configure 02-ip-vrf " + " \n\n")

with open(outfil, 'a') as outfile:
	outfile.write('\n# ***** end of 02-out-vrf' + '.cfg ***** \n\n')
script = "\n ***** program end ***** " + " "
print(script)
