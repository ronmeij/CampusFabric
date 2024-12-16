##########
# Name:			05-2-config-int-descrip-shut.py
# Vers:			4.04
# Date:			241211
# Auth:			Ronald Meijer
# Pyth:			version 3
# Dependency	05-cluster-port-allocation_RON-SAMPLE_v4.0
# ChangeLog		4.01	if RESERVED-SPB in interface name no further action is taken
#				4.02	added end statement
#				4.03	detects the csv file delimiter
#				4.04	added support for xlsx file type
#
##########
# Variables:
#	intrfc__id		interface id
#	intrfc__nm		interface name; "none" sets to default
#	snmp_lnktr		snmp link trap; "no" disables
#	transceivr		"shut" in column "Transceiver"
#	switch_id1		left switch name
#	switch_id2		right switch name
#
#	filename	list of the *.csv files in working directory
#	location	specifies working directory
#	fileset		filenames in the list filename
#	outfil		name of the out-file 
#	suffix		'.xlsx'
##########
# Run:      python3 config-int-descrip-shut.py
##########

import csv
import glob
# 'location' specifies working directory
# location = 'c:/test/temp/'
import os.path
import re
# the following may be required
# pip install openpyxl
import openpyxl
from openpyxl import load_workbook
directory = "."

def answer():
	answer = input("Press enter to continue")

print("\nThis script configures the interface 'name', admin-status and SNMP link trap for VOSS:")
print ("\nFile(s) found in local directory:")
fileset = [file for file in glob.glob("05*.xlsx", recursive=False)]

# Print the items with their corresponding indices
for idx, file in enumerate(fileset):
    print(f"    {idx+1}: {file}")

filename =""
#filename = input("\nPlease enter the filename: ")

while True:
	try:
		selected_index = int(input("Please select an index from the list of files: "))
	except ValueError:
		print("\n\t **** INFORM Please enter a valid integer for the index (a number please...)\n")
		continue
	else:
		# print (selected_index)
		# print (len(fileset)+1)
		if selected_index <= 0:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if selected_index >= len(fileset)+1:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if 1 <= selected_index <= len(fileset)+1:
			selected_file = fileset[selected_index-1]
			# Here you can open or use the file, for example:
			# Open the selected file
			filename = os.path.join(directory, selected_file)
			print("\n\t**** INFORM file found\n")
			print("\nColumn \t Header")
		break
with open (filename) as f:
	workbook = load_workbook(filename)
	# Select the active sheet (or use workbook['SheetName'])
	#sheet = workbook.active

	# Get all sheet names (tabs)
	sheet_names = workbook.sheetnames
	print("Available worksheets:", sheet_names)
  
	for sheet_name in sheet_names:
		print(f"\nReading data from sheet: {sheet_name}")
		sheet = workbook[sheet_name]
		# Select the active sheet (or use workbook['SheetName'])
		#sheet = workbook.active
    
		# Start processing from a specific row, e.g., row 5
		header_row_index = 5 
		
		#for header_row in sheet.iter_rows(min_row=header_row_index, values_only=True):
			#header_row = next(reader)
			
		header_row = [ cell.value for cell in sheet[header_row_index] if cell.value and str(cell.value).strip() ]
		
		for index, column_header in enumerate(header_row):
			print(index, "\t" ,column_header.strip())
			if int(0) == index:
				switch_id1 = column_header.strip()
			if int(12) == index:
				switch_id2 = column_header.strip()

		answer()
		intrfc__id = []
		intrfc__nm = []
		snmp_lnktr = []
		transceivr = []

		valid_rows = []

		#for row in reader:
		# Read table for SWITCH#1 starting with row 6 down and till column 12
		for row in sheet.iter_rows(min_row=header_row_index+1, min_col=1, max_col=12, values_only=True):
			if row[0] is not None:
				valid_rows.append(row)
				
		for row in valid_rows:
			try:
				intrfc__id.append(row[0] if row[0] is not None else "")
				intrfc__nm.append(row[3] if row[3] is not None else "")
				snmp_lnktr.append(row[4] if row[4] is not None else "")
				transceivr.append(row[5] if row[5] is not None else "")
			except IndexError:
				print("\n\n\t**** WARNING excel header not as expected!!!")
				quit()
				
				
		intrfc__id = [x.strip(" ") for x in intrfc__id]
		print(intrfc__id)
		intrfc__nm = [x.strip(" ") for x in intrfc__nm]
		print(intrfc__nm)
		snmp_lnktr = [x.strip(" ") for x in snmp_lnktr]
		snmp_lnktr = [x.lower() for x in snmp_lnktr]
		print(snmp_lnktr)
		transceivr = [x.strip(" ") for x in transceivr]
		transceivr = [x.lower() for x in transceivr]
		print(transceivr)
			
		answer()
		lenintrfc__id = len(intrfc__id)
		lenintrfc__nm = len(intrfc__nm)
		lensnmp_lnktr = len(snmp_lnktr)
		lentransceivr = len(transceivr)
		print(lenintrfc__id)
		print(lenintrfc__nm)
		print(lensnmp_lnktr)
		print(lentransceivr)
		print("Press enter to continue with the left-hand switch")
		answer()
			
		if not "NONE" in switch_id1:
			outfil = "05-2-out-config-int-descrip-shut-" + switch_id1 + ".cfg"
			print("\nThe name of the out-file is " + outfil + " \n")
			answer()
			with open(outfil, 'w') as outfile:
				outfile.write('\n# ***** start of ' + outfil + ' ***** ')
			with open(outfil, 'a') as outfile:
				outfile.write('\n\n')
			script = "rwa"
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("rwa\n")
			script = "rwa"
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("rwa\n")
			script = "enable " + " "
			print (script)
			with open(outfil, 'a') as outfile:
				outfile.write("enable " + "\n\n")
			script = "configure terminal " + " "
			print (script)
			with open(outfil, 'a') as outfile:
				outfile.write("configure terminal\n\n")

			# script = "# During this phase we set all interfaces as tagged - encapsulation dot1q " + " "
			# print (script)
			# with open(outfil, 'a') as outfile:
				# outfile.write("# During this phase we set all interfaces as tagged - encapsulation dot1q \n\n")

			for value in range(0,lenintrfc__id):
				if "RESERVED-SPB" in intrfc__nm[value]:
					script = "# No action is taken on a RESERVED-SPB interface " + intrfc__id[value] + " ** "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("\n# No action is taken on a RESERVED-SPB interface " + intrfc__id[value] + " ** ")
				else:
					if intrfc__id[value]:
						script = "\n\ninterface gigabitethernet " + intrfc__id[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("\n\ninterface gigabitethernet " + intrfc__id[value] + "\n")
						# script = "encapsulation dot1q " + " "
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("encapsulation dot1q " + "\n")		
						# script = "spoof-detect enable " + " "
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("spoof-detect enable " + "\n")
						if intrfc__nm[value]:
							if "none" in intrfc__nm[value]:
								script = 'default name ' + ' '
								print(script)
								with open(outfil, 'a') as outfile:
									outfile.write('default name ' + '\n')
							else:	
								script = 'name "' + intrfc__nm[value] + '" '
								print(script)
								with open(outfil, 'a') as outfile:
									outfile.write('name "' + intrfc__nm[value] + '"\n')
						if "shut" == transceivr[value]:
							script = "shutdown " + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("shutdown " + "\n")
						else:
							script = "no shutdown " + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("no shutdown " + "\n")
						if "no" == snmp_lnktr[value]:
							script = "no snmp trap link-status " + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("no snmp trap link-status " + "\n")
						script = "exit " + "\n"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("exit" + "\n")

		script = "\nend " + "\n\n"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\nend " + "\n\n")
		with open(outfil, 'a') as outfile:
			outfile.write('\n# ***** end of ' + outfil + ' ***** \n\n')

		print("\n#### Left-hand switch parsed succesfully\n\n")


		

		answer()

		intrfc__id = []
		intrfc__nm = []
		snmp_lnktr = []
		transceivr = []
		
		valid_rows = []

		#for row in reader:
		# Read table for SWITCH#1 starting with row 6 down and from column 14 till column 25
		for row in sheet.iter_rows(min_row=header_row_index+1, min_col=14, max_col=25, values_only=True):
			if row[0] is not None:
				valid_rows.append(row)
			
		for row in valid_rows:
			try:
				intrfc__id.append(row[0] if row[0] is not None else "")
				intrfc__nm.append(row[3] if row[3] is not None else "")
				snmp_lnktr.append(row[4] if row[4] is not None else "")
				transceivr.append(row[5] if row[5] is not None else "")
				#intrfc__id.append(row[13] if row[13] is not None else "")
				#intrfc__nm.append(row[16 if row[16] is not None else ""])
				#snmp_lnktr.append(row[17] if row[17] is not None else "")
				#transceivr.append(row[18] if row[18] is not None else "")
			except IndexError:
				print("\n\n\t**** WARNING excel header not as expected!!!")
				quit()
		intrfc__id = [x.strip(" ") for x in intrfc__id]
		print(intrfc__id)
		intrfc__nm = [x.strip(" ") for x in intrfc__nm]
		print(intrfc__nm)
		snmp_lnktr = [x.strip(" ") for x in snmp_lnktr]
		snmp_lnktr = [x.lower() for x in snmp_lnktr]
		print(snmp_lnktr)
		transceivr = [x.strip(" ") for x in transceivr]
		transceivr = [x.lower() for x in transceivr]
		print(transceivr)

		answer()
		lenintrfc__id = len(intrfc__id)
		lenintrfc__nm = len(intrfc__nm)
		lensnmp_lnktr = len(snmp_lnktr)
		lentransceivr = len(transceivr)
		print(lenintrfc__id)
		print(lenintrfc__nm)
		print(lensnmp_lnktr)
		print(lentransceivr)

		print("\nPress enter to continue with the right-hand switch")
		answer()

		if not "NONE" in switch_id2:
			outfil = "05-2-out-config-int-descrip-shut-" + switch_id2 + ".cfg"
			print("\nThe name of the out-file is " + outfil + " \n")
			answer()
			with open(outfil, 'w') as outfile:
				outfile.write('\n# ***** start of ' + outfil + ' ***** ')
			with open(outfil, 'a') as outfile:
				outfile.write('\n\n')
			script = "rwa"
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("rwa\n")
			script = "rwa"
			print(script)
			with open(outfil, 'a') as outfile:
				outfile.write("rwa\n")
			script = "enable " + " "
			print (script)
			with open(outfil, 'a') as outfile:
				outfile.write("enable " + "\n\n")
			script = "configure terminal " + " "
			print (script)
			with open(outfil, 'a') as outfile:
				outfile.write("configure terminal\n\n")

			# script = "# During this phase we set all interfaces as tagged - encapsulation dot1q " + " "
			# print (script)
			# with open(outfil, 'a') as outfile:
				# outfile.write("# During this phase we set all interfaces as tagged - encapsulation dot1q \n\n")

			for value in range(0,lenintrfc__id):
				if "RESERVED-SPB" in intrfc__nm[value]:
					script = "# No action is taken on a RESERVED-SPB interface " + intrfc__id[value] + " ** "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("\n# No action is taken on a RESERVED-SPB interface " + intrfc__id[value] + " ** ")
				else:
					if intrfc__id[value]:
						script = "\n\ninterface gigabitethernet " + intrfc__id[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("\n\ninterface gigabitethernet " + intrfc__id[value] + "\n")
						# script = "encapsulation dot1q " + " "
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("encapsulation dot1q " + "\n")		
						# script = "spoof-detect enable " + " "
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("spoof-detect enable " + "\n")
						if intrfc__nm[value]:
							if "none" in intrfc__nm[value]:
								script = 'default name ' + ' '
								print(script)
								with open(outfil, 'a') as outfile:
									outfile.write('default name ' + '\n')
							else:	
								script = 'name "' + intrfc__nm[value] + '" '
								print(script)
								with open(outfil, 'a') as outfile:
									outfile.write('name "' + intrfc__nm[value] + '"\n')
						if "shut" == transceivr[value]:
							script = "shutdown " + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("shutdown " + "\n")
						else:
							script = "no shutdown " + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("no shutdown " + "\n")
						if "no" == snmp_lnktr[value]:
							script = "no snmp trap link-status " + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("no snmp trap link-status " + "\n")
						script = "exit " + "\n"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("exit" + "\n")

		script = "\nend " + "\n\n"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\nend " + "\n\n")
		with open(outfil, 'a') as outfile:
			outfile.write('\n# ***** end of ' + outfil + ' ***** \n\n')

		print("\n#### Right-hand switch parsed succesfully\n\n")

		answer()


	script = "\n ***** program end ***** " + " "
	print(script)

