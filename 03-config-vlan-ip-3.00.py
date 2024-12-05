##########
# Name: 		03-config-vlan-ip.py
# Vers: 		3.00
# Date: 		241204
# Auth: 		Ronald Meijer
# Pyth: 		version 3
# Dependency:	03-vlan-id-name-isid-ip_RON-SAMPLE_v2.0
# Dependency:	openpyxl
# ChangeLog		2.90	added the option for simplified vIST
#				2.91	IPv6 support added (VRRP, RSMLT)
#						src-port-67 was added in dhcp-relay fwd-path as comment
#				2.92	added IGMP support for layer-2 VLANs
#				2.93	correction on syntax and added mc-config-lite support for grt
#				2.94	added option d for ip_adr_msk configures the ip interface to state-disabled
#				2.95	added file creation for enabling ip interfaces on vlans
#				2.96	use VRRP1 gateway as querier for mc-config-lite
#				2.97	corrections
#				2.98	added end statement
#				2.99	detects the csv file delimiter
#				3.00	added support for xlsx file type
##########
# Variables:
#	dhcp_rel_1 
#	dhcp_rel_2 
#	dhcp_rel_3 
#	dhcp_rel_4 
#	dvr____ena 
#	dvr___gtwy 
#	ip_dirbrdc 
#	igmp_proxy 
#	igmp_snoop 
#	ip_intadr1 	this value selects the 1st switch    (nocreate will not create the vlan)
#	ip_intadr2	this value selects the 2nd switch    (nocreate will not create the vlan)
#	ip_intadr3 	this value selects the 3rd switch    (nocreate will not create the vlan)
#	ip_intadr4	this value selects the 4th switch    (nocreate will not create the vlan)
#	ip_intadr5 	this value selects the 5th switch    (nocreate will not create the vlan)
#	ip_intadr6	this value selects the 6th switch    (nocreate will not create the vlan)
#	ip_intadr7 	this value selects the 7th switch    (nocreate will not create the vlan)
#	ip_intadr8	this value selects the 8th switch    (nocreate will not create the vlan)
#	ip_intadr9 	this value selects the 9th switch    (nocreate will not create the vlan)
#	ip_intadr0	this value selects the 10th switch   (nocreate will not create the vlan)
#	ip_intadrA	this value selects the 11th switch   (nocreate will not create the vlan)
#	ip_intadrB	this value selects the 12th switch   (nocreate will not create the vlan)
#	ip_adr_msk 
#	ip_mc__ena 	ip multicast enable in spbm on vlan
#	isid____id 	i-sid that maps to the vlan-id
#	ospf_passi 
#	rsmlt__ena 
#	rsmlt_edge 
#	vlan____id 	vlan-id
#	vlan__name 	vlan-name
#	vrf_____id 	vrf-id
#	vrf___name 	vrf-name
#	vrrp1__ena 
#	vrrp2__ena 
#	vrrp1_gtwy 
#	vrrp2_gtwy 
#	vrrp1___id 
#	vrrp2___id 
#	vrrp1_prio 
#	vrrp2_prio 
#	ist___type	selector for vIST or Simplified-vIST
#	ist___name	selected vIST type; used to determine the settings
#	state		used for interface state-disabled
#
#	filename	list of the *.csv files in working directory
#	location	specifies working directory
#	fileset		filenames in the list filename
#	outfil		name of the out-file
#	suffix		'.xlsx'
##########
# Run:      python3 config-vlan-ip.py
##########

import csv
import glob
# 'location' specifies working directory
# location = 'c:/test/temp/'
import os.path
import re

# the following may be required
# pip install openpyxl
import openpyxl
from openpyxl import load_workbook
directory = "."

def answer():
	answer = input("Press enter to continue")

print("\nThis script configures VLANs, IP, DHCP and SLPP for VOSS:")

# select uni-type
while True:
	try:
		ist___type = int(input("\n\nSelect the IST-type:\n 1 for Simplified-vIST (no SPBm)\n 2 for vIST  (SPBm)\n  : [2] ") or "2")
	except ValueError:
		print("Should be a value from 1 to 2")
		continue
	else:
		if ist___type <= 0:
			print("\nPlease select a value between 1 and 2")
			continue
		if ist___type >= 3:
			print("\nPlease select a value between 1 and 2")
			continue
		# uni-type number was succesfully parsed
		break
if ist___type == 1:
	ist___name = "SimpIST"
elif ist___type == 2:
	ist___name = "NormIST"
print("\n")
# uni-name was successfully set

print ("\n\nFile(s) found in local directory:")
fileset = [file for file in glob.glob("03*.xlsx", recursive=False)]

# Print the items with their corresponding indices
for idx, file in enumerate(fileset):
	print(f"    {idx+1}: {file}")

filename =""

while True:
	try:
		selected_index = int(input("Please select an index from the list of files: "))
	except ValueError:
		print("\n\t **** INFORM Please enter a valid integer for the index (a number please...)\n")
		continue
	else:
		# print (selected_index)
		# print (len(fileset)+1)
		if selected_index <= 0:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if selected_index >= len(fileset)+1:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if 1 <= selected_index <= len(fileset)+1:
			selected_file = fileset[selected_index-1]
			# Here you can open or use the file, for example:
			# Open the selected file
			filename = os.path.join(directory, selected_file)
			print("\n\t**** INFORM file found\n")
			print("\nColumn \t Header")
		break
with open (filename) as f:
	workbook = load_workbook(filename)
	# Select the active sheet (or use workbook['SheetName'])
	sheet = workbook.active
	# Start processing from a specific row, e.g., row 3
	header_row_index = 3 
	#for header_row in sheet.iter_rows(min_row=header_row_index, values_only=True):
		#header_row = next(reader)
	header_row = [ cell.value for cell in sheet[header_row_index] if cell.value and str(cell.value).strip() ]
	for index, column_header in enumerate(header_row):
		print(index, "\t", column_header.strip())

	print("\n\t**** INFORM The file " + filename + " was selected\n")
	answer()

	for index, column_header in enumerate(header_row):
		print(index, "\t" ,column_header.strip())
		if int(28) == index:
			switch_id1 = column_header.strip()
		if int(29) == index:
			switch_id2 = column_header.strip()
		if int(30) == index:
			switch_id3 = column_header.strip()
		if int(31) == index:
			switch_id4 = column_header.strip()
		if int(32) == index:
			switch_id5 = column_header.strip()
		if int(33) == index:
			switch_id6 = column_header.strip()
		if int(34) == index:
			switch_id7 = column_header.strip()
		if int(35) == index:
			switch_id8 = column_header.strip()
		if int(36) == index:
			switch_id9 = column_header.strip()
		if int(37) == index:
			switch_id0 = column_header.strip()
		if int(38) == index:
			switch_idA = column_header.strip()
		if int(39) == index:
			switch_idB = column_header.strip()
	answer()
	dhcp_rel_1 = []
	dhcp_rel_2 = []
	dhcp_rel_3 = []
	dhcp_rel_4 = []
	dvr____ena = []
	dvr___gtwy = []
	ip_dirbrdc = []
	igmp_proxy = []
	igmp_snoop = []
	ip_intadr1 = []
	ip_intadr2 = []
	ip_intadr3 = []
	ip_intadr4 = []
	ip_intadr5 = []
	ip_intadr6 = []
	ip_intadr7 = []
	ip_intadr8 = []
	ip_intadr9 = []
	ip_intadr0 = []
	ip_intadrA = []
	ip_intadrB = []
	ip_adr_msk = []
	ip_mc__ena = []
	isid____id = []
	ospf_passi = []
	rsmlt__ena = []
	rsmlt_edge = []
	vlan____id = []
	vlan__name = []
	vrf_____id = []
	vrf___name = []
	vrrp1__ena = []
	vrrp2__ena = []
	vrrp1_gtwy = []
	vrrp2_gtwy = []
	vrrp1___id = []
	vrrp2___id = []
	vrrp1_prio = []
	vrrp2_prio = []

	valid_rows = []

	#for row in reader:
	for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
		if row[0] is not None:
			valid_rows.append(row)

	print('XXXXXXX Nb valid rows=', len(valid_rows))

	for row in valid_rows:
		try:
			dhcp_rel_1.append(row[19] if row[19] is not None else "")
			dhcp_rel_2.append(row[20] if row[20] is not None else "")
			dhcp_rel_3.append(row[21] if row[21] is not None else "")
			dhcp_rel_4.append(row[22] if row[22] is not None else "")
			dvr____ena.append(row[6] if row[6] is not None else "")
			dvr___gtwy.append(row[7] if row[7] is not None else "")
			ip_dirbrdc.append(row[26] if row[26] is not None else "")
			igmp_proxy.append(row[24] if row[24] is not None else "")
			igmp_snoop.append(row[25] if row[25] is not None else "")
			ip_intadr1.append(row[28] if row[28] is not None else "")
			ip_intadr2.append(row[29] if row[29] is not None else "")
			ip_intadr3.append(row[30] if row[30] is not None else "")
			ip_intadr4.append(row[31] if row[31] is not None else "")
			ip_intadr5.append(row[32] if row[32] is not None else "")
			ip_intadr6.append(row[33] if row[33] is not None else "")
			ip_intadr7.append(row[34] if row[34] is not None else "")
			ip_intadr8.append(row[35] if row[35] is not None else "")
			ip_intadr9.append(row[36] if row[36] is not None else "")
			ip_intadr0.append(row[37] if row[37] is not None else "")
			ip_intadrA.append(row[38] if row[38] is not None else "")
			ip_intadrB.append(row[39] if row[39] is not None else "")
			ip_adr_msk.append(row[27] if row[27] is not None else "")
			ip_mc__ena.append(row[23] if row[23] is not None else "")
			isid____id.append(row[2] if row[2] is not None else "")
			ospf_passi.append(row[18] if row[18] is not None else "")
			rsmlt__ena.append(row[8] if row[8] is not None else "")
			rsmlt_edge.append(row[9] if row[9] is not None else "")
			vlan____id.append(row[0] if row[0] is not None else "")
			vlan__name.append(row[1] if row[1] is not None else "")
			vrf_____id.append(row[5] if row[5] is not None else "")
			vrf___name.append(row[4] if row[4] is not None else "")
			vrrp1__ena.append(row[13] if row[13] is not None else "")
			vrrp2__ena.append(row[17] if row[17] is not None else "")
			vrrp1_gtwy.append(row[12] if row[12] is not None else "")
			vrrp2_gtwy.append(row[16] if row[16] is not None else "")
			vrrp1___id.append(row[10] if row[10] is not None else "")
			vrrp2___id.append(row[14] if row[14] is not None else "")
			vrrp1_prio.append(row[11] if row[11] is not None else "")
			vrrp2_prio.append(row[15] if row[15] is not None else "")
		except IndexError:
			print("\n\n\t**** WARNING excel header not as expected!!!")
			quit()
	print("\n\n**** List data:")
	print("\t**** dhcp-relay agent 1")
	dhcp_rel_1 = [x.encode("ascii", "ignore") for x in dhcp_rel_1]
	dhcp_rel_1 = [x.decode() for x in dhcp_rel_1]
	dhcp_rel_1 = [x.strip(" ") for x in dhcp_rel_1]
	print(dhcp_rel_1)
	print("\t**** dhcp-relay agent 2")
	dhcp_rel_2 = [x.encode("ascii", "ignore") for x in dhcp_rel_2]
	dhcp_rel_2 = [x.decode() for x in dhcp_rel_2]
	dhcp_rel_2 = [x.strip(" ") for x in dhcp_rel_2]
	print(dhcp_rel_2)
	print("\t**** dhcp-relay agent 3")
	dhcp_rel_3 = [x.encode("ascii", "ignore") for x in dhcp_rel_3]
	dhcp_rel_3 = [x.decode() for x in dhcp_rel_3]
	dhcp_rel_3 = [x.strip(" ") for x in dhcp_rel_3]
	print(dhcp_rel_3)
	print("\t**** dhcp-relay agent 4")
	dhcp_rel_4 = [x.encode("ascii", "ignore") for x in dhcp_rel_4]
	dhcp_rel_4 = [x.decode() for x in dhcp_rel_4]
	dhcp_rel_4 = [x.strip(" ") for x in dhcp_rel_4]
	print(dhcp_rel_4)
	print("\t**** dvr enable")
	dvr____ena = [x.encode("ascii", "ignore") for x in dvr____ena]
	dvr____ena = [x.decode() for x in dvr____ena]
	dvr____ena = [x.strip(" ") for x in dvr____ena]
	dvr____ena = [x.lower() for x in dvr____ena]
	print(dvr____ena)
	print("\t**** dvr gateway")
	dvr___gtwy = [x.encode("ascii", "ignore") for x in dvr___gtwy]
	dvr___gtwy = [x.decode() for x in dvr___gtwy]
	dvr___gtwy = [x.strip(" ") for x in dvr___gtwy]
	print(dvr___gtwy)
	print("\t**** ip directed broadcast")
	ip_dirbrdc = [x.encode("ascii", "ignore") for x in ip_dirbrdc]
	ip_dirbrdc = [x.decode() for x in ip_dirbrdc]
	ip_dirbrdc = [x.strip(" ") for x in ip_dirbrdc]
	ip_dirbrdc = [x.lower() for x in ip_dirbrdc]
	print(ip_dirbrdc)
	print("\t**** igmp proxy")
	igmp_proxy = [x.encode("ascii", "ignore") for x in igmp_proxy]
	igmp_proxy = [x.decode() for x in igmp_proxy]
	igmp_proxy = [x.strip(" ") for x in igmp_proxy]
	igmp_proxy = [x.lower() for x in igmp_proxy]
	print(igmp_proxy)
	print("\t**** igmp snoop")
	igmp_snoop = [x.encode("ascii", "ignore") for x in igmp_snoop]
	igmp_snoop = [x.decode() for x in igmp_snoop]
	igmp_snoop = [x.strip(" ") for x in igmp_snoop]
	igmp_snoop = [x.lower() for x in igmp_snoop]
	print(igmp_snoop)
	print("\t**** ip interface address switch 1")
	ip_intadr1 = [x.encode("ascii", "ignore") for x in ip_intadr1]
	ip_intadr1 = [x.decode() for x in ip_intadr1]
	ip_intadr1 = [x.strip(" ") for x in ip_intadr1]
	print(ip_intadr1)
	print("\t**** ip interface address switch 2")
	ip_intadr2 = [x.encode("ascii", "ignore") for x in ip_intadr2]
	ip_intadr2 = [x.decode() for x in ip_intadr2]
	ip_intadr2 = [x.strip(" ") for x in ip_intadr2]
	print(ip_intadr2)
	print("\t**** ip interface address switch 3")
	ip_intadr3 = [x.encode("ascii", "ignore") for x in ip_intadr3]
	ip_intadr3 = [x.decode() for x in ip_intadr3]
	ip_intadr3 = [x.strip(" ") for x in ip_intadr3]
	print(ip_intadr3)
	print("\t**** ip interface address switch 4")
	ip_intadr4 = [x.encode("ascii", "ignore") for x in ip_intadr4]
	ip_intadr4 = [x.decode() for x in ip_intadr4]
	ip_intadr4 = [x.strip(" ") for x in ip_intadr4]
	print(ip_intadr4)
	print("\t**** ip interface address switch 5")
	ip_intadr5 = [x.encode("ascii", "ignore") for x in ip_intadr5]
	ip_intadr5 = [x.decode() for x in ip_intadr5]
	ip_intadr5 = [x.strip(" ") for x in ip_intadr5]
	print(ip_intadr5)
	print("\t**** ip interface address switch 6")
	ip_intadr6 = [x.encode("ascii", "ignore") for x in ip_intadr6]
	ip_intadr6 = [x.decode() for x in ip_intadr6]
	ip_intadr6 = [x.strip(" ") for x in ip_intadr6]
	print(ip_intadr6)
	print("\t**** ip interface address switch 7")
	ip_intadr7 = [x.encode("ascii", "ignore") for x in ip_intadr7]
	ip_intadr7 = [x.decode() for x in ip_intadr7]
	ip_intadr7 = [x.strip(" ") for x in ip_intadr7]
	print(ip_intadr7)
	print("\t**** ip interface address switch 8")
	ip_intadr8 = [x.encode("ascii", "ignore") for x in ip_intadr8]
	ip_intadr8 = [x.decode() for x in ip_intadr8]
	ip_intadr8 = [x.strip(" ") for x in ip_intadr8]
	print(ip_intadr8)
	print("\t**** ip interface address switch 9")
	ip_intadr9 = [x.encode("ascii", "ignore") for x in ip_intadr9]
	ip_intadr9 = [x.decode() for x in ip_intadr9]
	ip_intadr9 = [x.strip(" ") for x in ip_intadr9]
	print(ip_intadr9)
	print("\t**** ip interface address switch 10")
	ip_intadr0 = [x.encode("ascii", "ignore") for x in ip_intadr0]
	ip_intadr0 = [x.decode() for x in ip_intadr0]
	ip_intadr0 = [x.strip(" ") for x in ip_intadr0]
	print(ip_intadr0)
	print("\t**** ip interface address switch 11")
	ip_intadrA = [x.encode("ascii", "ignore") for x in ip_intadrA]
	ip_intadrA = [x.decode() for x in ip_intadrA]
	ip_intadrA = [x.strip(" ") for x in ip_intadrA]
	print(ip_intadrA)
	print("\t**** ip interface address switch 12")
	ip_intadrB = [x.encode("ascii", "ignore") for x in ip_intadrB]
	ip_intadrB = [x.decode() for x in ip_intadrB]
	ip_intadrB = [x.strip(" ") for x in ip_intadrB]
	print(ip_intadrB)
	print("\t**** ip netmask")
	ip_adr_msk = [x.encode("ascii", "ignore") for x in ip_adr_msk]
	ip_adr_msk = [x.decode() for x in ip_adr_msk]
	ip_adr_msk = [x.strip(" ") for x in ip_adr_msk]
	ip_adr_msk = [x.lower() for x in ip_adr_msk]
	# ip_adr_msk = [x.isspace() for x in ip_adr_msk]
	print(ip_adr_msk)
	print("\t**** ip multicast enable")
	ip_mc__ena = [x.encode("ascii", "ignore") for x in ip_mc__ena]
	ip_mc__ena = [x.decode() for x in ip_mc__ena]
	ip_mc__ena = [x.strip(" ") for x in ip_mc__ena]
	ip_mc__ena = [x.lower() for x in ip_mc__ena]
	print(ip_mc__ena)
	print("\t**** l2vsn i-sid")
	isid____id = [x.encode("ascii", "ignore") for x in isid____id]
	isid____id = [x.decode() for x in isid____id]
	isid____id = [x.strip(" ") for x in isid____id]
	print(isid____id)
	print("\t**** ip ospf passive interface")
	ospf_passi = [x.encode("ascii", "ignore") for x in ospf_passi]
	ospf_passi = [x.decode() for x in ospf_passi]
	ospf_passi = [x.strip(" ") for x in ospf_passi]
	ospf_passi = [x.lower() for x in ospf_passi]
	print(ospf_passi)
	print("\t**** ip rsmlt enable")
	rsmlt__ena = [x.encode("ascii", "ignore") for x in rsmlt__ena]
	rsmlt__ena = [x.decode() for x in rsmlt__ena]
	rsmlt__ena = [x.strip(" ") for x in rsmlt__ena]
	rsmlt__ena = [x.lower() for x in rsmlt__ena]
	print(rsmlt__ena)
	print("\t**** ip rsmlt edge enable")
	rsmlt_edge = [x.encode("ascii", "ignore") for x in rsmlt_edge]
	rsmlt_edge = [x.decode() for x in rsmlt_edge]
	rsmlt_edge = [x.strip(" ") for x in rsmlt_edge]
	rsmlt_edge = [x.lower() for x in rsmlt_edge]
	print(rsmlt_edge)
	print("\t**** vlan id")
	# vlan____id = [x.encode("ascii", "ignore") for x in vlan____id]
	# vlan____id = [x.decode() for x in vlan____id]
	# vlan____id = [x.strip(" ") for x in vlan____id]
	print(vlan____id)
	print("\t**** vlan name")
	vlan__name = [x.encode("ascii", "ignore") for x in vlan__name]
	vlan__name = [x.decode() for x in vlan__name]
	vlan__name = [x.strip(" ") for x in vlan__name]
	vlan__name = [x.strip("\t") for x in vlan__name]
	print(vlan__name)
	print("\t**** vrf id (0=grt)")
	vrf_____id = [x.encode("ascii", "ignore") for x in vrf_____id]
	vrf_____id = [x.decode() for x in vrf_____id]
	vrf_____id = [x.strip(" ") for x in vrf_____id]
	print(vrf_____id)
	print("\t**** vrf name")
	vrf___name = [x.encode("ascii", "ignore") for x in vrf___name]
	vrf___name = [x.decode() for x in vrf___name]
	vrf___name = [x.strip(" ") for x in vrf___name]
	print(vrf___name)
	print("\t**** ip vrrp first instance enable")
	vrrp1__ena = [x.encode("ascii", "ignore") for x in vrrp1__ena]
	vrrp1__ena = [x.decode() for x in vrrp1__ena]
	vrrp1__ena = [x.strip(" ") for x in vrrp1__ena]
	vrrp1__ena = [x.lower() for x in vrrp1__ena]
	print(vrrp1__ena)
	print("\t**** ip vrrp second instance enable")
	vrrp2__ena = [x.encode("ascii", "ignore") for x in vrrp2__ena]
	vrrp2__ena = [x.decode() for x in vrrp2__ena]
	vrrp2__ena = [x.strip(" ") for x in vrrp2__ena]
	vrrp2__ena = [x.lower() for x in vrrp2__ena]
	print(vrrp2__ena)
	print("\t**** ip vrrp gateway 1")
	vrrp1_gtwy = [x.encode("ascii", "ignore") for x in vrrp1_gtwy]
	vrrp1_gtwy = [x.decode() for x in vrrp1_gtwy]
	vrrp1_gtwy = [x.strip(" ") for x in vrrp1_gtwy]
	print(vrrp1_gtwy)
	print("\t**** ip vrrp gateway 2")
	vrrp2_gtwy = [x.encode("ascii", "ignore") for x in vrrp2_gtwy]
	vrrp2_gtwy = [x.decode() for x in vrrp2_gtwy]
	vrrp2_gtwy = [x.strip(" ") for x in vrrp2_gtwy]
	print(vrrp2_gtwy)
	print("\t**** ip vrrp vrid 1")
	vrrp1___id = [x.encode("ascii", "ignore") for x in vrrp1___id]
	vrrp1___id = [x.decode() for x in vrrp1___id]
	vrrp1___id = [x.strip(" ") for x in vrrp1___id]
	print(vrrp1___id)
	print("\t**** ip vrrp vrid 2")
	vrrp2___id = [x.encode("ascii", "ignore") for x in vrrp2___id]
	vrrp2___id = [x.decode() for x in vrrp2___id]
	vrrp2___id = [x.strip(" ") for x in vrrp2___id]
	print(vrrp2___id)
	print("\t**** ip vrrp priorities instance 1")
	vrrp1_prio = [x.encode("ascii", "ignore") for x in vrrp1_prio]
	vrrp1_prio = [x.decode() for x in vrrp1_prio]
	vrrp1_prio = [x.strip(" ") for x in vrrp1_prio]
	print(vrrp1_prio)
	print("\t**** ip vrrp priorities instance 2")
	vrrp2_prio = [x.encode("ascii", "ignore") for x in vrrp2_prio]
	vrrp2_prio = [x.decode() for x in vrrp2_prio]
	vrrp2_prio = [x.strip(" ") for x in vrrp2_prio]
	print(vrrp2_prio)
answer()
lendhcp_rel_1 = len(dhcp_rel_1)
lendhcp_rel_2 = len(dhcp_rel_2)
lendhcp_rel_3 = len(dhcp_rel_3)
lendhcp_rel_4 = len(dhcp_rel_4)
lendvr____ena = len(dvr____ena)
lendvr___gtwy = len(dvr___gtwy)
lenip_dirbrdc = len(ip_dirbrdc)
lenigmp_proxy = len(igmp_proxy)
lenigmp_snoop = len(igmp_snoop)
lenip_intadr1 = len(ip_intadr1)
lenip_intadr2 = len(ip_intadr2)
lenip_intadr3 = len(ip_intadr3)
lenip_intadr4 = len(ip_intadr4)
lenip_intadr5 = len(ip_intadr5)
lenip_intadr6 = len(ip_intadr6)
lenip_intadr7 = len(ip_intadr7)
lenip_intadr8 = len(ip_intadr8)
lenip_intadr9 = len(ip_intadr9)
lenip_intadr0 = len(ip_intadr0)
lenip_intadrA = len(ip_intadrA)
lenip_intadrB = len(ip_intadrB)
lenip_adr_msk = len(ip_adr_msk)
lenip_mc__ena = len(ip_mc__ena)
lenisid____id = len(isid____id)
lenospf_passi = len(ospf_passi)
lenrsmlt__ena = len(rsmlt__ena)
lenrsmlt_edge = len(rsmlt_edge)
lenvlan____id = len(vlan____id)
lenvlan__name = len(vlan__name)
lenvrf_____id = len(vrf_____id)
lenvrf___name = len(vrf___name)
lenvrrp1__ena = len(vrrp1__ena)
lenvrrp2__ena = len(vrrp2__ena)
lenvrrp1_gtwy = len(vrrp1_gtwy)
lenvrrp2_gtwy = len(vrrp2_gtwy)
lenvrrp1___id = len(vrrp1___id)
lenvrrp2___id = len(vrrp2___id)
lenvrrp1_prio = len(vrrp1_prio)
lenvrrp2_prio = len(vrrp2_prio)
print("\n\n**** Number of list entries (should be all the same):")
print(lendhcp_rel_1)
print(lendhcp_rel_2)
print(lendhcp_rel_3)
print(lendhcp_rel_4)
print(lendvr____ena)
print(lendvr___gtwy)
print(lenip_dirbrdc)
print(lenigmp_proxy)
print(lenigmp_snoop)
print(lenip_intadr1)
print(lenip_intadr2)
print(lenip_intadr3)
print(lenip_intadr4)
print(lenip_intadr5)
print(lenip_intadr6)
print(lenip_intadr7)
print(lenip_intadr8)
print(lenip_intadr9)
print(lenip_intadr0)
print(lenip_intadrA)
print(lenip_intadrB)
print(lenip_adr_msk)
print(lenip_mc__ena)
print(lenisid____id)
print(lenospf_passi)
print(lenrsmlt__ena)
print(lenrsmlt_edge)
print(lenvlan____id)
print(lenvlan__name)
print(lenvrf_____id)
print(lenvrf___name)
print(lenvrrp1__ena)
print(lenvrrp2__ena)
print(lenvrrp1_gtwy)
print(lenvrrp2_gtwy)
print(lenvrrp1___id)
print(lenvrrp2___id)
print(lenvrrp1_prio)
print(lenvrrp2_prio)

print("\n #### Configuration dictionary looks like   ####")
print("\n #### 'NONE' as switch name in column header skips column ####\n\n\n")
def switch_fnc(switch_idx, ip_intadrx):
	if not "NONE" in switch_idx:
		outfil = "03-out-config-vlan-ip-" + switch_idx + ".cfg"
		print("\nThe name of the out-file is " + outfil + " \n")
		answer()
		with open(outfil, 'w') as outfile:
			outfile.write('\n# ***** start of ' + outfil + ' ***** ')
		with open(outfil, 'a') as outfile:
			outfile.write('\n# ***** ' + ist___name + '-Type was selected ***** ')
		with open(outfil, 'a') as outfile:
			outfile.write('\n\n')
		script = "rwa"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("rwa\n")
		script = "rwa"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("rwa\n")
		script = "enable " + " "
		print (script)
		with open(outfil, 'a') as outfile:
			outfile.write("enable " + "\n\n")
		script = "configure terminal " + " "
		print (script)
		with open(outfil, 'a') as outfile:
			outfile.write("configure terminal\n\n")

		script = "\n\n# STEP 1 - Create VLANs and enable SLPP on VLANs"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\n# STEP 1 - Create VLANs and enable SLPP on VLANs \n\n")
		for value in range(0,lenvlan____id):
			if "nocreate" in ip_intadrx[value]:
				script = '# Not create vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' on this switch (nocreate) '
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write('# Not create vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' on this switch (nocreate) \n\n') 
			else:
				script = 'vlan create ' + str(vlan____id[value]) + ' name "' + vlan__name[value] + '" type port-mstprstp 0 '
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write('vlan create ' + str(vlan____id[value]) + ' name "' + vlan__name[value] + '" type port-mstprstp 0 \n')
				if ist___name == "NormIST":
					if not isid____id[value]:
						script = '# Error no i-sid defined for VLAN ' + vlan____id[value] + ' ' + vlan__name[value] + ' '
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write('# Error no i-sid defined for VLAN ' + vlan____id[value] + ' ' + vlan__name[value] + ' \n\n') 
					else:
						script = "vlan i-sid " + str(vlan____id[value]) + " " + isid____id[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("vlan i-sid " + str(vlan____id[value]) + " " + isid____id[value] + " \n\n")
				script = "slpp vid " + str(vlan____id[value]) + " "
				print (script)
				with open(outfil, 'a') as outfile:
					outfile.write("slpp vid " + str(vlan____id[value]) + " \n\n\n")

		script = "\n\n# STEP 2 - Create IP interface on VLAN"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\n# STEP 2 - Create IP interface on VLAN \n\n")
		for value in range(0,lenvlan____id):
			if "nocreate" in ip_intadrx[value] or "mc-config-lite" in ip_intadrx[value]:
				if "nocreate" in ip_intadrx[value]:
					script = '# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created on this switch, no IP interface'
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write('\n# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created on this switch, no IP interface \n\n') 
				if "mc-config-lite" in ip_intadrx[value]:
					script = "interface vlan " + str(vlan____id[value]) + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("\ninterface vlan " + str(vlan____id[value]) + " \n")
					# if int(vrf_____id[value]) >= 1:
						# script = "vrf " + vrf___name[value] + " "
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("vrf " + vrf___name[value] + " \n")
					# else int(vrf_____id[value]) >= 1:
						# script = "vrf " + vrf___name[value] + " "
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("vrf " + vrf___name[value] + " \n")
					script = "mvpn-isid 0" + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("mvpn-isid 0" + " \n")
					script = "ip spb-multicast enable" + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("ip spb-multicast enable" + " \n")
					script = "# ip igmp routed-spb-querier-addr " + vrrp1_gtwy[value] + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("# ip igmp routed-spb-querier-addr " + vrrp1_gtwy[value] + " \n")
					script = "exit" + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("exit" + " \n\n")

			elif ip_intadrx[value]:
				state = ''
				if "d" in ip_adr_msk[value]:
					# ip_adr_msk[value] = (ip_adr_msk[value]).strip('d')
					# print(ip_adr_msk[value])
					state = ' state-disabled'
					# print(state)
					script = '# Note that the IP address on VLAN ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' will be created with state-disabled '
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write('\n# Note that the IP address on VLAN ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' will be created with state-disabled ' + ' \n')

				script = "interface vlan " + str(vlan____id[value]) + " "
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write("\ninterface vlan " + str(vlan____id[value]) + " \n")
				if int(vrf_____id[value]) >= 1:
					script = "vrf " + vrf___name[value] + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("vrf " + vrf___name[value] + " \n")

				if dvr___gtwy[value]:
					if ist___name == "SimpIST":
						script = '# Error DVR-IP cannot be configured with Simplified vIST for VLAN ' + str(vlan____id[value]) + ' '
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write('# Error DVR-IP cannot be configured with Simplified vIST for VLAN ' + str(vlan____id[value]) + ' \n\n')
					else:
						if dvr___gtwy[value] == ip_intadrx[value]:
							script = "ip address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + " dvr-one-ip " + state + " " 
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + " dvr-one-ip" + state + "\n")
						else:
							script = "ip address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + state + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + state + " \n")
							script = "dvr gw-ipv4 " + dvr___gtwy[value] + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("dvr gw-ipv4 " + dvr___gtwy[value] + " \n")
					if "yes" in dvr____ena[value]:
						if ist___name == "SimpIST":
							script = '# Error DVR-IP cannot be enabled with Simplified vIST for VLAN ' + str(vlan____id[value]) + ' '
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('# Error DVR-IP cannot be enabled with Simplified vIST for VLAN ' + str(vlan____id[value]) + ' \n\n')
						else:
							if not dvr___gtwy[value]:
								script = '# Error DVR-IP cannot be enabled without DVR gateway address for VLAN ' + str(vlan____id[value]) + ' '
								print(script)
								with open(outfil, 'a') as outfile:
									outfile.write('# Error DVR-IP cannot be enabled without DVR gateway address for VLAN ' + str(vlan____id[value]) + ' \n\n')
							else:
								script = "dvr enable"
								print(script)
								with open(outfil, 'a') as outfile:
									outfile.write("dvr enable \n")
				else:
					if ":" in ip_intadrx[value]:
						script = 'ipv6 interface enable '
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write('ipv6 interface enable \n') 
						script = "ipv6 interface address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + state + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 interface address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + state + " \n")
					else:
						script = "ip address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + state + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip address " + ip_intadrx[value] + (ip_adr_msk[value]).strip('d') + state + " \n")

				if "yes" in ip_mc__ena[value]:
					script = "ip spb-multicast enable "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("ip spb-multicast enable \n")
	
				if "yes" in ip_dirbrdc[value]:
					script = "ip directed-broadcast enable "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("ip directed-broadcast enable \n")
	
				if "yes" in ospf_passi[value]:
					script = "ip ospf network passive"
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("ip ospf network passive \n")
					script = "ip ospf enable"
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("ip ospf enable \n")
					script = "y"
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("y \n")
	
				if "yes" in igmp_proxy[value]:
					if not "yes" in dvr____ena[value]:
						script = "ip igmp proxy"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip igmp proxy \n")
	
				if "yes" in igmp_snoop[value]:
					if not "yes" in dvr____ena[value]:
						script = "ip igmp snooping"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip igmp snooping \n")
					else:
						script = "ip spb-multicast enable "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip spb-multicast enable \n")

				if dhcp_rel_1[value]:
					script = "ip dhcp-relay"
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("ip dhcp-relay \n")
					script = "ip dhcp-relay mode dhcp"
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("ip dhcp-relay mode dhcp \n")

				if "yes" in rsmlt__ena[value]:
					if ":" in  ip_intadrx[value]:
						if int(vrf_____id[value]) >= 1:
							script = '# Error IPv6 RSMLT is not supported under VRF for VLAN ' + str(vlan____id[value]) + ' '
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('# Error IPv6 RSMLT is not supported under VRF for VLAN ' + str(vlan____id[value]) + ' \n\n')
							quit()
						else: 
							script = "ipv6 rsmlt"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ipv6 rsmlt \n")
							if "yes" in rsmlt_edge[value]:
								script = "ipv6 rsmlt holdup-timer 9999"
								print(script)
								with open(outfil, 'a') as outfile:
									outfile.write("ipv6 rsmlt holdup-timer 9999 \n")
					else:
						script = "ip rsmlt"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip rsmlt \n")
						# script = "ip rsmlt holddown-timer 90"
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("ip rsmlt holddown-timer 90 \n")
						# script = "ip rsmlt holdup-timer 120"
						# print(script)
						# with open(outfil, 'a') as outfile:
							# outfile.write("ip rsmlt holdup-timer 120 \n")
						if "yes" in rsmlt_edge[value]:
							script = "ip rsmlt holdup-timer 9999"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip rsmlt holdup-timer 9999 \n")

				if vrrp1_gtwy[value]:
					if ":" in vrrp1_gtwy[value]:
						script = "ipv6 vrrp address " + vrrp1___id[value] + " link-local fe80:0:0:0:0:0:" +  vrrp1___id[value] + ":1"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp address " + vrrp1___id[value] + " link-local fe80:0:0:0:0:0:" +  vrrp1___id[value] + ":1\n")
						script = "ipv6 vrrp address " + vrrp1___id[value] + " global " +  vrrp1_gtwy[value] + "/" + (ip_adr_msk[value]).strip('d') + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp address " + vrrp1___id[value] + " global " +  vrrp1_gtwy[value] + "/" + (ip_adr_msk[value]).strip('d') + "\n")
						script = "ipv6 vrrp " + vrrp1___id[value] + " backup-master enable "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp " + vrrp1___id[value] + " backup-master enable \n")
						script = "ipv6 vrrp " + vrrp1___id[value] + " priority " + vrrp1_prio[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp " + vrrp1___id[value] + " priority " + vrrp1_prio[value] + " \n")
					else:
						script = "ip vrrp version 3"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp version 3 \n")
						script = "ip vrrp address " + vrrp1___id[value] + " " + vrrp1_gtwy[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp address " + vrrp1___id[value] + " " + vrrp1_gtwy[value] + " \n")
						script = "ip vrrp " + vrrp1___id[value] + " backup-master enable "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp " + vrrp1___id[value] + " backup-master enable \n")
						script = "ip vrrp " + vrrp1___id[value] + " priority " + vrrp1_prio[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp " + vrrp1___id[value] + " priority " + vrrp1_prio[value] + " \n")
				if "yes" in vrrp1__ena[value]:
					if ":" in vrrp1_gtwy[value]:
						script = "ipv6 vrrp " + vrrp1___id[value] + " enable"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp " + vrrp1___id[value] + " enable \n")
					else:
						script = "ip vrrp " + vrrp1___id[value] + " enable"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp " + vrrp1___id[value] + " enable \n")

				if vrrp2_gtwy[value]:
					if ":" in vrrp2_gtwy[value]:
						script = "ipv6 vrrp address " + vrrp2___id[value] + " link-local fe80:0:0:0:0:0:" +  vrrp2___id[value] + ":1"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp address " + vrrp2___id[value] + " link-local fe80:0:0:0:0:0:" +  vrrp2___id[value] + ":1\n")
						script = "ipv6 vrrp address " + vrrp2___id[value] + " global " +  vrrp2_gtwy[value] + "/" + (ip_adr_msk[value]).strip('d') + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp address " + vrrp2___id[value] + " global " +  vrrp2_gtwy[value] + "/" + (ip_adr_msk[value]).strip('d') + "\n")
						script = "ipv6 vrrp " + vrrp2___id[value] + " backup-master enable "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp " + vrrp2___id[value] + " backup-master enable \n")
						script = "ipv6 vrrp " + vrrp2___id[value] + " priority " + vrrp2_prio[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp " + vrrp2___id[value] + " priority " + vrrp2_prio[value] + " \n")
					else:
						script = "ip vrrp address " + vrrp2___id[value] + " " + vrrp2_gtwy[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp address " + vrrp2___id[value] + " " + vrrp2_gtwy[value] + " \n")
						script = "ip vrrp " + vrrp2___id[value] + " backup-master enable "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp " + vrrp2___id[value] + " backup-master enable \n")
						script = "ip vrrp " + vrrp2___id[value] + " priority " + vrrp2_prio[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp " + vrrp2___id[value] + " priority " + vrrp2_prio[value] + " \n")
				if "yes" in vrrp2__ena[value]:
					if ":" in vrrp2_gtwy[value]:
						script = "ipv6 vrrp " + vrrp2___id[value] + " enable"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ipv6 vrrp " + vrrp2___id[value] + " enable \n")
					else:
						script = "ip vrrp " + vrrp2___id[value] + " enable"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip vrrp " + vrrp2___id[value] + " enable \n")
				else:
					script = "exit "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("exit " + "\n")
							
			elif "yes" in igmp_snoop[value]:
				script = "interface vlan " + str(vlan____id[value]) + " "
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write("\ninterface vlan " + str(vlan____id[value]) + " \n")
				if int(vrf_____id[value]) >= 1:
					script = "vrf " + vrf___name[value] + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("vrf " + vrf___name[value] + " \n")
				script = "ip igmp snooping"
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write("ip igmp snooping \n")
				script = "# ip igmp snoop-querier-addr x.x.x.x"
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write("# ip igmp snoop-querier-addr x.x.x.x \n")
				script = "# ip igmp snoop-querier"
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write("# ip igmp snoop-querier \n")

				script = "exit"
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write("exit \n\n")
	
			if "nocreate" in ip_intadrx[value] or "mc-config-lite" in ip_intadrx[value]:
				script = '# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created on this switch, no DHCP-relay'
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write('# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created on this switch, no DHCP-relay \n\n') 
			else:
				if dhcp_rel_1[value]:
					if ip_intadrx[value]:
						if int(vrf_____id[value]) >= 1:
							script = "router vrf " + vrf___name[value] + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("router vrf " + vrf___name[value] + " \n")
						script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " "
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " \n")
						script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " enable"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " enable\n")
						script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " mode dhcp"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " mode dhcp\n")
						script = "# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " src-port-67"
						print(script)
						with open(outfil, 'a') as outfile:
							outfile.write("# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_1[value] + " src-port-67\n")
						if dhcp_rel_2[value]:
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " \n")
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " enable"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " enable\n")
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " mode dhcp"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " mode dhcp\n")
							script = "# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " src-port-67"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_2[value] + " src-port-67\n")
						if dhcp_rel_3[value]:
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " \n")
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " enable"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " enable\n")
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " mode dhcp"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " mode dhcp\n")
							script = "# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " src-port-67"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_3[value] + " src-port-67\n")
						if dhcp_rel_4[value]:
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " \n")
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " enable"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " enable\n")
							script = "ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " mode dhcp"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " mode dhcp\n")
							script = "# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " src-port-67"
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("# ip dhcp-relay fwd-path " + ip_intadrx[value] + " " + dhcp_rel_4[value] + " src-port-67\n")
						if int(vrf_____id[value]) >= 1:
							script = "exit" + " "
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write("exit" + " \n\n")

		script = "\nend " + " "
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\nend " + "\n\n")
		with open(outfil, 'a') as outfile:
			outfile.write('\n# ***** end of ' + outfil + ' ***** \n\n')

		answer()
		if not "NONE" in switch_idx:
			outfil = "03-out-enable-ip-" + switch_idx + ".cfg"
			print("\nThe name of the out-file is " + outfil + " \n")
			answer()
			with open(outfil, 'w') as outfile:
				outfile.write('\n# ***** start of ' + outfil + ' ***** ')
			with open(outfil, 'a') as outfile:
				outfile.write('\n# ***** ' + ist___name + '-Type was selected ***** ')
			with open(outfil, 'a') as outfile:
				outfile.write('\n\n')
			script = "configure terminal " + " "
			print (script)
			with open(outfil, 'a') as outfile:
				outfile.write("configure terminal\n\n")

		script = "\n\n# STEP 1 - Enable IP interface on created VLANs"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\n# STEP 1 - Enable IP interface on created VLANs \n\n")
		for value in range(0,lenvlan____id):
			if "nocreate" in ip_intadrx[value] or "mc-config-lite" in ip_intadrx[value]:
				if "nocreate" in ip_intadrx[value]:
					script = '# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created on this switch, no IP interface'
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write('\n# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created on this switch, no IP interface \n\n') 
				if "mc-config-lite" in ip_intadrx[value]:
					script = '# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created (mc-config-lite) on this switch, no IP interface'
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write('\n# The vlan ' + str(vlan____id[value]) + ' ' + vlan__name[value] + ' is not created (mc-config-lite) on this switch, no IP interface \n\n') 
			elif ip_intadrx[value]:
				script = "interface vlan " + str(vlan____id[value]) + " "
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write("\ninterface vlan " + str(vlan____id[value]) + " \n")
				script = 'ip interface enable '
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write('ip interface enable \n') 
				script = 'exit '
				print(script)
				with open(outfil, 'a') as outfile:
					outfile.write('exit \n\n') 
		script = "\nend " + " "
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\nend " + "\n\n")
		with open(outfil, 'a') as outfile:
			outfile.write('\n# ***** end of ' + outfil + ' ***** \n\n')
		answer()

# build dictionary

switch_lib = {}

switch_lib = {
	switch_id1 : ip_intadr1,
	switch_id2 : ip_intadr2,
	switch_id3 : ip_intadr3,
	switch_id4 : ip_intadr4,
	switch_id5 : ip_intadr5,
	switch_id6 : ip_intadr6,
	switch_id7 : ip_intadr7,
	switch_id8 : ip_intadr8,
	switch_id9 : ip_intadr9,
	switch_id0 : ip_intadr0,
	switch_idA : ip_intadrA,
	switch_idB : ip_intadrB,
	}

# print(switch_lib)

answer()

for key, value in switch_lib.items():
	print("\nSwitch: " + key + "\tInterface IP address: ")
	for values in value:
		print("\t\t\t" + values)
	switch_fnc(key, value)

script = "\n ***** program end ***** " + " "
print(script)

