##########
# Name: display-header.py
# Vers: 1.00
# Date: 241204
# Auth: Ronald Meijer
# Pyth: version 3
##########
# Dependency:	openpyxl
# ChangeLog:	1.00
#				1.00	reads xlsx file type
##########
# Variables:
#	filename	list of the '*.csv' files in working directory
#	location	specifies working directory
#	fileset		filenames in the list filename
#	suffix		'.xlsx'
##########
# Run:      python3 000-display-header-xlsx-1.xx.py
##########

import csv
import glob
# 'location' specifies working directory
# location = 'c:/test/temp/'
import os.path
import re
# the following may be required
# pip install openpyxl
import openpyxl
from openpyxl import load_workbook
directory = "."

def answer():
	answer = input("Press enter to continue")

print("\nThis script reads the header of a XLSX file")
print ("\nFile(s) found in local directory:")
fileset = [file for file in glob.glob("*.xlsx", recursive=False)]

# Print the items with their corresponding indices
for idx, file in enumerate(fileset):
	print(f"    {idx+1}: {file}")

filename =""

while True:
	try:
		selected_index = int(input("Please select an index from the list of files: "))
	except ValueError:
		print("\n\t **** INFORM Please enter a valid integer for the index (a number please...)\n")
		continue
	else:
		# print (selected_index)
		# print (len(fileset)+1)
		if selected_index <= 0:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if selected_index >= len(fileset)+1:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if 1 <= selected_index <= len(fileset)+1:
			selected_file = fileset[selected_index-1]
			# Here you can open or use the file, for example:
			# Open the selected file
			filename = os.path.join(directory, selected_file)
			print("\n\t**** INFORM file found\n")
			print("\nColumn \t Header")
		break
with open (filename) as f:
	workbook = load_workbook(filename)
	# Select the active sheet (or use workbook['SheetName'])
	sheet = workbook.active
	# Start processing from a specific row, e.g., row 3
	header_row_index = 3 
	#for header_row in sheet.iter_rows(min_row=header_row_index, values_only=True):
		#header_row = next(reader)
	header_row = [ cell.value for cell in sheet[header_row_index] if cell.value and str(cell.value).strip() ]
	for index, column_header in enumerate(header_row):
		print(index, "\t", column_header.strip())

	print("\n\tThe file " + filename + " was selected\n")
	answer()

script = "\n ***** program end ***** " + " "
print(script)

