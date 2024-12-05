##########
# Name: 		05-1-config-dict-vlan-isid.py
# Vers: 		1.62
# Date: 		241204
# Auth: 		Ronald Meijer
# Pyth: 		version 3
# Dependency:	03-vlan-id-name-isid-ip_RON-SAMPLE_v2.0
# Dependency:	openpyxl
# ChangeLog:	1.61	detects the csv file delimiter
#				1.62	added support for xlsx file type
##########
# Variables:
#	vlan____id	vlan-id
#	isid____id	i-sid that maps to the vlan-id
#	vlan__elan	vlan-id to elan-id mapping dictionary
#
#	filename	list of the *.csv files in working directory
#	location	specifies working directory
#	fileset		filenames in the list filename
#	outfilename	name of the out-file 
#	suffix		'.xlsx'
##########
# Run:      python3 config-dict-vlan-isid.py
##########

import csv
import glob
# 'location' specifies working directory
# location = 'c:/test/temp/'
import os.path
import re

# the following may be required
# pip install openpyxl
import openpyxl
from openpyxl import load_workbook
directory = "."

def answer():
	answer = input("Press enter to continue")

print("\nThis script populates the VLAN to I-SID dictionary used in 05-x-config-flex-uni-x.xx.py")
print ("\nFile(s) found in local directory:")
fileset = [file for file in glob.glob("03*.xlsx", recursive=False)]

# Print the items with their corresponding indices
for idx, file in enumerate(fileset):
	print(f"    {idx+1}: {file}")

filename =""

while True:
	try:
		selected_index = int(input("Please select an index from the list of files: "))
	except ValueError:
		print("\n\t **** INFORM Please enter a valid integer for the index (a number please...)\n")
		continue
	else:
		# print (selected_index)
		# print (len(fileset)+1)
		if selected_index <= 0:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if selected_index >= len(fileset)+1:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if 1 <= selected_index <= len(fileset)+1:
			selected_file = fileset[selected_index-1]
			# Here you can open or use the file, for example:
			# Open the selected file
			filename = os.path.join(directory, selected_file)
			print("\n\t**** INFORM file found\n")
			print("\nColumn \t Header")
		break
with open (filename) as f:
	workbook = load_workbook(filename)
	# Select the active sheet (or use workbook['SheetName'])
	sheet = workbook.active
	# Start processing from a specific row, e.g., row 3
	header_row_index = 3 
	#for header_row in sheet.iter_rows(min_row=header_row_index, values_only=True):
		#header_row = next(reader)
	header_row = [ cell.value for cell in sheet[header_row_index] if cell.value and str(cell.value).strip() ]
	for index, column_header in enumerate(header_row):
		print(index, "\t", column_header.strip())

	print("\n\t**** INFORM The file " + filename + " was selected\n")
	answer()

	vlan____id = []
	isid____id = []

	valid_rows = []

	#for row in reader:
	for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
		if row[0] is not None:
			valid_rows.append(row)

	print('XXXXXXX Nb valid rows=', len(valid_rows))

	for row in valid_rows:
		try:
			vlan____id.append(row[0] if row[0] is not None else "")
			isid____id.append(row[2] if row[2] is not None else "")
		except IndexError:
			print("\n\n\t**** WARNING Excel header not as expected!!!")
			quit()
print("\n\n**** List data from Excel file:")

print("\t**** VLANid:")
# vlan____id = [x.strip(" ") for x in vlan____id]
print(vlan____id)
print("\t**** I-SID:")
# isid____id = [x.strip(" ") for x in isid____id]
print(isid____id)
len_vlan____id = len(vlan____id)
len_isid____id = len(isid____id)
print(len_vlan____id)
print(len_isid____id)
answer()
outfilename = "05-1-out-dict-vlan-isid"
print("\nThe name of the out-file is " + outfilename + ".txt \n")
answer()

vlan__isid = {}

# build dictionary
for value in range(0,len_vlan____id):
	if vlan____id[value]:
		vlan__isid[vlan____id[value]] = isid____id[value]
# print number of vlans in excel file
print("Number of entries in Excel file: " + str(len_vlan____id))
# print the number of keys in dictionary
print("Number of entries in dictionary: " + (str(len(vlan__isid.keys()))))
len_vlan__isid = len(vlan__isid.keys())

answer()

for key, value in vlan__isid.items():
	print("VLAN: " + str(key) + "\tI-SID: " + str(value))

# convert key to integer
new_vlan_isid = {int(old_key): val for old_key, val in vlan__isid.items()}

# sort dictionary as vlan_isid_sorted
vlan__isid_sorted = {}
for i in sorted(new_vlan_isid):
   vlan__isid_sorted[i] = new_vlan_isid[i]
# print(vlan__isid_sorted)

# write file
print("\nSorted output:")
with open(outfilename + '.txt', 'w') as outfile:
	outfile.write("{")
for key, value in vlan__isid_sorted.items():
	print("VLAN: " + str(key) + "\tI-SID: " + value)
	with open(outfilename + '.txt', 'a') as outfile:
		outfile.write("\n    " + str(key) + " : " + value + ",")
with open(outfilename + '.txt', 'a') as outfile:
	outfile.write("\n    }" + "\n")

# Finding duplicated vlan ids
if len_vlan__isid < len_vlan____id:
	print("\n\t**** WARNING duplicated VLAN id(s) found: ")
# Printing original list
# print ("Original list is : " + str(vlan____id))

# convert list to integers
for i in range(0, len(vlan____id)):
	vlan____id[i] = int(vlan____id[i])

# Printing modified list 
# print ("Modified list is : " + str(vlan____id))
# Searching duplicate intergers in list
def repeat(x):
	_size = len(x)
	repeated = []
	for i in range(_size):
		k = i + 1
		for j in range(k, _size):
			if x[i] == x[j] and x[i] not in repeated:
				repeated.append(x[i])
	return repeated

# Print duplicates
# print("\t\t\tThe following VLAN id(s) are duplicated: " + str(repeat(vlan____id)))
len_repeat = len(repeat(vlan____id))
# print(len_repeat)
for vlan in range(0,len_repeat):
	print("\t\t Duplicated VLAN: " + str(repeat(vlan____id)[vlan]) + " ")

# Finding duplicate isid values in a dictionary

# printing initial_dictionary
# print("Initial_dictionary :", str(vlan__isid_sorted))

# finding duplicate values in dictionary
rev_dict = {}
  
for key, value in vlan__isid_sorted.items():
    rev_dict.setdefault(value, set()).add(key)

result = [key for key, values in rev_dict.items() if len(values) > 1]

# printing result
if str(result) != "[]":
	print("\n\n\t**** WARNING duplicate I-SIDs found: ")
	# print("\n\n\t**** WARNING duplicate I-SIDs found: " + str(result) + " Please check information!!!")

def getKeysByValue(dictOfElements, valueToFind):
	listOfKeys = list()
	listOfItems = dictOfElements.items()
	for item  in listOfItems:
		if item[1] == valueToFind:
			listOfKeys.append(item[0])
	return listOfKeys

for isid in result:
	listOfKeys = getKeysByValue(vlan__isid_sorted, isid)
	print("\t\tVLANs with an I-SID equal to " + str(isid) + ": ")
	for key  in listOfKeys:
		print("\t\t\t VLAN: " + str(key))

answer()

script = "\n ***** program end ***** " + " "
print(script)

