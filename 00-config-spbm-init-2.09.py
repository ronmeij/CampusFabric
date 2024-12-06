##########
# Name: 		00-config-spbm-init.py
# Vers: 		2.09
# Date: 		241204
# Auth: 		Ronald Meijer
# Pyth: 		version 3
# Dependency:	01_spbm_param_detail_TEMPLATE_v2.1 or 01_spbm_param_detail_TEMPLATE_v3.0
# Dependency:	openpyxl
# ChangeLog:	1.52	added #force-topology-ip under mgmt oob
#				1.53	simplified vIST support
#							uses promptname as file-name
#							uses out-of-band ip address
#						corrected ssh for max-session (ssh default enabled)
#				1.54	no reboot required anymore for dvr leaf function (from 8.5.0.0 onwards)
#				2.00	new excel lay-out
#						prompt name is used as file name for out-file
#						only segmented management is supported
#				2.01	added more ssh commands (ssh bouncing too fast)
#						added #sys vim-speed 10000 as remark
#				2.02	corrected vim-speed (added yes)
#						corrected order for dhcp client
#				2.03	added the command show io resources nlb to gain time after ssh
#				2.04	added a dummy route entry for oob mgmt interface
#						added a warning message concerning factory defaults before script
#				2.05	changed some screen messages
#				2.06	try to enable ssh twice, in case fails first time
#				2.07	cleanup ssh 8 sessions
#				2.08	detects the csv file delimiter
#				2.09	added support for xlsx file type
##########
# provides Out-of-Band management interface
# ssh max-sessions 8
##########
# Variables:
#	dvr___func 	dvr function; ctrl, leaf or none
#	ist___name	selected vIST type; used to determine the settings
#	ist___type	selector for vIST or Simplified-vIST
#	ob_mgmt_ip 	out-of-band management IP address in format x.x.x.x/yy
#	promptname 	switch prompt name
#
#	filename	filename of the selected .csv file
#	location	specifies working directory, not used (current directory)
#	fileset		list of the *.csv files in working directory
#	out_file	name of the out-file
#	suffix		'.xlsx' 
##########
# Run:      python3 00-config-spbm-init-x.xx.py
##########
import csv
import glob
# 'location' specifies working directory
# location = 'c:/test/temp/'
import os.path
import re
# the following may be required
# pip install openpyxl
import openpyxl
from openpyxl import load_workbook
directory = "."

def answer():
	answer = input("Press enter to continue")

print("\nThis script configures initial settings for VOSS:")
print("(boot 'flags', 'Out-of_Band' interface, etc...)")
print("\nThe file should be copied into the console of the factory defaulted VOSS switch\n")
print("\n\n### Make sure to 'factory-default' the switch before running the script: ###")
print("\tenable")
print("\tconfigure terminal \n\tboot config flags factorydefaults \n\texit \n\tboot -y")
print("### Failing FACTORY DEFAULTS will not remove ZTP settings !!! ###")

# select uni-type
while True:
	try:
		ist___type = int(input("\n\nSelect the IST-type:\n 1 for Simplified-vIST setup (NO SPBm, for PIM deployments)\n 2 for vIST setup (SPBm, for Campus Fabric deployments)\n  : [2] ") or "2")
	except ValueError:
		print("Should be a value from 1 to 2")
		continue
	else:
		if ist___type <= 0:
			print("\nPlease select a value between 1 and 2")
			continue
		if ist___type >= 3:
			print("\nPlease select a value between 1 and 2")
			continue
		# SPBm or Simplified vIST mode was succesfully parsed
		break
if ist___type == 1:
	ist___name = "SimpIST"
elif ist___type == 2:
	ist___name = "NormIST"
print("\n")
# ist-name was successfully set
print ("\nFile(s) found in local directory:")
fileset = [file for file in glob.glob("01*.xlsx", recursive=False)]

# Print the items with their corresponding indices
for idx, file in enumerate(fileset):
	print(f"    {idx+1}: {file}")

filename =""

while True:
	try:
		selected_index = int(input("Please select an index from the list of files: "))
	except ValueError:
		print("\n\t **** INFORM Please enter a valid integer for the index (a number please...)\n")
		continue
	else:
		# print (selected_index)
		# print (len(fileset)+1)
		if selected_index <= 0:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if selected_index >= len(fileset)+1:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if 1 <= selected_index <= len(fileset)+1:
			selected_file = fileset[selected_index-1]
			# Here you can open or use the file, for example:
			# Open the selected file
			filename = os.path.join(directory, selected_file)
			print("\n\t**** INFORM file found\n")
			print("\nColumn \t Header")
		break
with open (filename) as f:
	workbook = load_workbook(filename)
	# Select the active sheet (or use workbook['SheetName'])
	sheet = workbook.active
	# Start processing from a specific row, e.g., row 3
	header_row_index = 3 
	#for header_row in sheet.iter_rows(min_row=header_row_index, values_only=True):
		#header_row = next(reader)
	header_row = [ cell.value for cell in sheet[header_row_index] if cell.value and str(cell.value).strip() ]
	for index, column_header in enumerate(header_row):
		print(index, "\t", column_header.strip())

	print("\n\tThe file " + filename + " was selected\n")
	answer()

	promptname = []
	ob_mgmt_ip = []
	dvr___func = []

	valid_rows = []
	
	#for row in reader:
	for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
		if row[0] is not None:
			valid_rows.append(row)
			
	print('\n\n\tXXXXXXX Nb valid rows=', len(valid_rows))

	for row in valid_rows:
		try:
			promptname.append(row[0] if row[0] is not None else "")
			ob_mgmt_ip.append(row[22] if row[22] is not None else "")
			dvr___func.append(row[23] if row[23] is not None else "")
		except IndexError:
			print("\n\n\t**** WARNING Excel header not as expected!!!")
			quit()
	print("\n\n**** List data from Excel file:")

	print("\n\t**** PromptName:")
	promptname = [x.strip(" ") for x in promptname]
	print(promptname)
	print("\n\t**** Out-of-Band mgmt ip address:")
	ob_mgmt_ip = [x.strip(" ") for x in ob_mgmt_ip]
	print(ob_mgmt_ip)
	print("\n\t**** DVR switch function (none, ctrl, leaf):")
	dvr___func = [x.strip(" ") for x in dvr___func]
	dvr___func = [x.lower() for x in dvr___func]
	print(dvr___func)

	print("\nInput data was correctly parsed from the Excel file\n")
	answer()
	lenpromptname = len(promptname)
	lenob_mgmt_ip = len(ob_mgmt_ip)
	lendvr___func = len(dvr___func)
	print(lenpromptname)
	print(lenob_mgmt_ip)
	print(lendvr___func)
	print("\n**** Number of switches in Excel file (all numbers above must be equal):\n")
	answer()
for value in range(0,lenpromptname):

	if promptname[value] == "":
		print("\n\n\t**** DATA error: switch must have a prompt name -> Excel file line " + str(value + 4) + " ")
		quit()
	else:
		out_file = "00-out-spbm-" + promptname[value] + ".cfg"
		print("\nThe name of the out-file is " + out_file + " \n")
		with open(out_file, 'w') as outfile:
			outfile.write('\n# ***** start of 00-out-spbm-' + promptname[value] + '.cfg ***** ')
		with open(out_file, 'a') as outfile:
			outfile.write('\n\n')

	script = "### Make sure to 'factory-default' the switch before running this script:"
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("### Make sure to 'factory-default' the switch before running this script: \n")
	script = "#       enable"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("#       enable \n")
	script = "#       configure terminal"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("#       configure terminal \n")
	script = "#       boot config flags factorydefaults "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("#       boot config flags factorydefaults \n")
	script = "#       exit"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("#       exit \n")
	script = "#       boot -y"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("#       boot -y \n")
	script = "### Failing FACTORY DEFAULTS will not remove ZTP settings !!! "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("### Failing FACTORY DEFAULTS will not remove ZTP settings !!! \n\n\n")

	script = "# user/password: " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("# user/password \n")
	script = "rwa"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("rwa\n")
	script = "rwa"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("rwa\n")
	script = "rwa"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("rwa\n")
	script = "rwa"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("rwa\n\n")
	script = "enable " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("enable " + "\n\n")

	script = "configure terminal " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("configure terminal\n\n")

	script = "boot config flags sshd " + " "
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("boot config flags sshd " + " \n")
	script = "boot config flags ftpd " + " "
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("boot config flags ftpd " + " \n")
	script = "# boot config flags telnetd " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("# boot config flags telnetd " + " \n\n")

	script = "# REMINDER set vim-speed to correct speed - 10/25G defaults to 25G" + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("# REMINDER set vim-speed to correct speed - 10/25G defaults to 25G \n")
	script = "# sys vim-speed 10000"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("# sys vim-speed 10000\n")
	script = "# y"
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("# y \n\n")

	script = "application " + " "
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("application " + " \n")
	script = "no iqagent enable " + " "
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("no iqagent enable " + " \n")
	script = "exit " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("exit " + " \n\n")

	script = "# no ssh " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("# no ssh " + " \n")
	script = "# ssh max-sessions 8 " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("# ssh max-sessions 8 " + " \n\n")
	script = "# enable ssh "
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("# enable ssh \n")
	script = "ssh " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("ssh " + " \n\n")

	script = "no mgmt dhcp-client" + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("no mgmt dhcp-client" + " \n\n")

	# script = "# trying to enable ssh again (in case it failed before)... "
	# print(script)
	# with open(out_file, 'a') as outfile:
		# outfile.write("# trying to enable ssh again (in case it failed before)... \n\n")
	# script = "ssh " + " "
	# print (script)
	# with open(out_file, 'a') as outfile:
		# outfile.write("ssh " + " \n\n")

	if ob_mgmt_ip[value]:
		script = "mgmt oob " + " "
		print(script)
		with open(out_file, 'a') as outfile:
			outfile.write("mgmt oob " + " \n")
		script = "ip address " + ob_mgmt_ip[value] + " "
		print(script)
		with open(out_file, 'a') as outfile:
			outfile.write("ip address " + ob_mgmt_ip[value] + " \n")
		script = "# ip route 0.0.0.0/0 next-hop y.y.y.y weight 300 "
		print(script)
		with open(out_file, 'a') as outfile:
			outfile.write("# ip route 0.0.0.0/0 next-hop y.y.y.y weight 300 \n")
		script = "enable " + " "
		print (script)
		with open(out_file, 'a') as outfile:
			outfile.write("enable " + " \n")
		script = "# force-topology-ip " + " "
		print (script)
		with open(out_file, 'a') as outfile:
			outfile.write("# force-topology-ip " + " \n")
		script = "exit " + " "
		print (script)
		with open(out_file, 'a') as outfile:
			outfile.write("exit " + " \n\n")

	if "leaf" in dvr___func[value]:
		script = "boot config flags dvr-leaf-mode" + " "
		print(script)
		with open(out_file, 'a') as outfile:
			outfile.write("boot config flags dvr-leaf-mode" + " \n")
		script = "y"
		print(script)
		with open(out_file, 'a') as outfile:
			outfile.write("y\n")

	if ist___name == "SimpIST":
		script = "no boot config flags spbm-config-mode" + " "
		print(script)
		with open(out_file, 'a') as outfile:
			outfile.write("no boot config flags spbm-config-mode" + " \n")

	script = "\nend " + " "
	print(script)
	with open(out_file, 'a') as outfile:
		outfile.write("\nend " + "\n\n")
	script = "save config " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("save config " + " \n\n")
	script = "backup configure 00-spbm-init " + " "
	print (script)
	with open(out_file, 'a') as outfile:
		outfile.write("backup configure 00-spbm-init " + " \n\n")

	if ist___name == "SimpIST":
		script = "boot -y " + " "
		print(script)
		with open(out_file, 'a') as outfile:
			outfile.write("boot -y " + " \n")

	with open(out_file, 'a') as outfile:
		outfile.write('\n# ***** end of 00-out-spbm-' + promptname[value] + '.cfg ***** \n\n')
script = "\n	**** program end **** " + " "
print(script)
