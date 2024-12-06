##########
# Name: 04-config-vrf-loopback.py
# Vers: 2.89
# Date: 241205
# Auth: Ronald Meijer
# Pyth: version 3
# Dependency:	04-config-vrf-loopback_RON-SAMPLE_v2.1
# Dependency:	openpyxl
# changeLog:	2.86	added IPv6 support
#				2.87	added end statement
#				2.88	detects the csv file delimiter
#				2.89	added support for xlsx file type
##########
# Variables:
#	lpbkint_id 	loopback interface id (2-256; 1 is used for ISIS loopback)
#	vrf___name 	vrf name
#	lpbck_ip_1 	vrf ip loopback address (switch column 1)
#	lpbck_ip_2 	vrf ip loopback address (switch column 2)
#	lpbck_ip_3 	vrf ip loopback address (switch column 3)
#	lpbck_ip_4 	vrf ip loopback address (switch column 4)
#	lpbck_ip_5 	vrf ip loopback address (switch column 5)
#	lpbck_ip_6 	vrf ip loopback address (switch column 6)
#	lpbck_ip_7 	vrf ip loopback address (switch column 7)
#	lpbck_ip_8 	vrf ip loopback address (switch column 8)
#	lpbck_ip_9 	vrf ip loopback address (switch column 9)
#	lpbck_ip_0 	vrf ip loopback address (switch column 10)
#	lpbck_ip_A 	vrf ip loopback address (switch column 11)
#	lpbck_ip_B 	vrf ip loopback address (switch column 12)
#	lpbck_name 	name for the loopback interface (CLIP_vrfname)
#
#	filename	list of the *.csv files in working directory
#	location	specifies working directory
#	fileset		filenames in the list filename
#	outfil		name of the out-file
#	suffix		'.xlsx' 
##########
# Run:      python3 config-vrf-loopback.py
##########

import csv
import glob
# 'location' specifies working directory
# location = 'c:/test/temp/'
import os.path
import re
# the following may be required
# pip install openpyxl
import openpyxl
from openpyxl import load_workbook
directory = "."

def answer():
	answer = input("Press enter to continue")

print("\nThis script configures IPv4 and IPv6 loopback addresses for GRT and VRFs in VOSS:")
print ("\nFile(s) found in local directory:")
fileset = [file for file in glob.glob("04*.xlsx", recursive=False)]

# Print the items with their corresponding indices
for idx, file in enumerate(fileset):
	print(f"    {idx+1}: {file}")

filename =""

while True:
	try:
		selected_index = int(input("Please select an index from the list of files: "))
	except ValueError:
		print("\n\t **** INFORM Please enter a valid integer for the index (a number please...)\n")
		continue
	else:
		# print (selected_index)
		# print (len(fileset)+1)
		if selected_index <= 0:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if selected_index >= len(fileset)+1:
			print ("\n\t**** INFORM Cannot open this file, please enter a valid integer for the index\n")
			continue
		if 1 <= selected_index <= len(fileset)+1:
			selected_file = fileset[selected_index-1]
			# Here you can open or use the file, for example:
			# Open the selected file
			filename = os.path.join(directory, selected_file)
			print("\n\t**** INFORM file found\n")
			print("\nColumn \t Header")
		break
with open (filename) as f:
	workbook = load_workbook(filename)
	# Select the active sheet (or use workbook['SheetName'])
	sheet = workbook.active
	# Start processing from a specific row, e.g., row 3
	header_row_index = 3 
	#for header_row in sheet.iter_rows(min_row=header_row_index, values_only=True):
		#header_row = next(reader)
	header_row = [ cell.value for cell in sheet[header_row_index] if cell.value and str(cell.value).strip() ]
	# for index, column_header in enumerate(header_row):
		# print(index, "\t", column_header.strip())

	for index, column_header in enumerate(header_row):
		print(index, "\t" ,column_header.strip())
		if int(3) == index:
			switch_id1 = column_header.strip()
		if int(4) == index:
			switch_id2 = column_header.strip()
		if int(5) == index:
			switch_id3 = column_header.strip()
		if int(6) == index:
			switch_id4 = column_header.strip()
		if int(7) == index:
			switch_id5 = column_header.strip()
		if int(8) == index:
			switch_id6 = column_header.strip()
		if int(9) == index:
			switch_id7 = column_header.strip()
		if int(10) == index:
			switch_id8 = column_header.strip()
		if int(11) == index:
			switch_id9 = column_header.strip()
		if int(12) == index:
			switch_id0 = column_header.strip()
		if int(13) == index:
			switch_idA = column_header.strip()
		if int(14) == index:
			switch_idB = column_header.strip()

	print("\n\t**** INFORM The file " + filename + " was selected\n")
	answer()

	lpbkint_id = []
	vrf___name = []
	lpbck_ip_1 = []
	lpbck_ip_2 = []
	lpbck_ip_3 = []
	lpbck_ip_4 = []
	lpbck_ip_5 = []
	lpbck_ip_6 = []
	lpbck_ip_7 = []
	lpbck_ip_8 = []
	lpbck_ip_9 = []
	lpbck_ip_0 = []
	lpbck_ip_A = []
	lpbck_ip_B = []
	lpbck_name = []

	valid_rows = []

	#for row in reader:
	for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
		if row[0] is not None:
			valid_rows.append(row)

	print('\n\tXXXXXXX Nb valid rows=', len(valid_rows))

	for row in valid_rows:
		try:
			lpbkint_id.append(row[2] if row[2] is not None else "")
			vrf___name.append(row[1] if row[1] is not None else "")
			lpbck_ip_1.append(row[3] if row[3] is not None else "")
			lpbck_ip_2.append(row[4] if row[4] is not None else "")
			lpbck_ip_3.append(row[5] if row[5] is not None else "")
			lpbck_ip_4.append(row[6] if row[6] is not None else "")
			lpbck_ip_5.append(row[7] if row[7] is not None else "")
			lpbck_ip_6.append(row[8] if row[8] is not None else "")
			lpbck_ip_7.append(row[9] if row[9] is not None else "")
			lpbck_ip_8.append(row[10] if row[10] is not None else "")
			lpbck_ip_9.append(row[11] if row[11] is not None else "")
			lpbck_ip_0.append(row[12] if row[12] is not None else "")
			lpbck_ip_A.append(row[13] if row[13] is not None else "")
			lpbck_ip_B.append(row[14] if row[14] is not None else "")
			lpbck_name.append(row[1] if row[1] is not None else "")
		except IndexError:
			print("\n\n\t**** WARNING excel header not as expected!!!")
			quit()
	print("\n\n**** List data:")
	print("\n\t**** loopback interface id (2-256 and 1 is used for ISIS loopback):")
	lpbkint_id = [x.strip(" ") for x in lpbkint_id]
	print(lpbkint_id)
	print("\t**** vrf name:")
	vrf___name = [x.strip(" ") for x in vrf___name]
	print(vrf___name)
	print("\t**** loopback ip switch 1:")
	lpbck_ip_1 = [x.strip(" ") for x in lpbck_ip_1]
	print(lpbck_ip_1)
	print("\t**** loopback ip switch 2:")
	lpbck_ip_2 = [x.strip(" ") for x in lpbck_ip_2]
	print(lpbck_ip_2)
	print("\t**** loopback ip switch 3:")
	lpbck_ip_3 = [x.strip(" ") for x in lpbck_ip_3]
	print(lpbck_ip_3)
	print("\t**** loopback ip switch 4:")
	lpbck_ip_4 = [x.strip(" ") for x in lpbck_ip_4]
	print(lpbck_ip_4)
	print("\t**** loopback ip switch 5:")
	lpbck_ip_5 = [x.strip(" ") for x in lpbck_ip_5]
	print(lpbck_ip_5)
	print("\t**** loopback ip switch 6:")
	lpbck_ip_6 = [x.strip(" ") for x in lpbck_ip_6]
	print(lpbck_ip_6)
	print("\t**** loopback ip switch 7:")
	lpbck_ip_7 = [x.strip(" ") for x in lpbck_ip_7]
	print(lpbck_ip_7)
	print("\t**** loopback ip switch 8:")
	lpbck_ip_8 = [x.strip(" ") for x in lpbck_ip_8]
	print(lpbck_ip_8)
	print("\t**** loopback ip switch 9:")
	lpbck_ip_9 = [x.strip(" ") for x in lpbck_ip_9]
	print(lpbck_ip_9)
	print("\t**** loopback ip switch 10:")
	lpbck_ip_0 = [x.strip(" ") for x in lpbck_ip_0]
	print(lpbck_ip_0)
	print("\t**** loopback ip switch 11:")
	lpbck_ip_A = [x.strip(" ") for x in lpbck_ip_A]
	print(lpbck_ip_A)
	print("\t**** loopback ip switch 12:")
	lpbck_ip_B = [x.strip(" ") for x in lpbck_ip_B]
	print(lpbck_ip_B)
	print("\t**** loopback interface name:")
	lpbck_name = [x.strip(" ") for x in lpbck_name]
	print(lpbck_name)
	print("\nInput data was correctly parsed from the Excel file\n")
answer()
lenlpbkint_id = len(lpbkint_id)
lenvrf___name = len(vrf___name)
lenlpbck_ip_1 = len(lpbck_ip_1)
lenlpbck_ip_2 = len(lpbck_ip_2)
lenlpbck_ip_3 = len(lpbck_ip_3)
lenlpbck_ip_4 = len(lpbck_ip_4)
lenlpbck_ip_5 = len(lpbck_ip_5)
lenlpbck_ip_6 = len(lpbck_ip_6)
lenlpbck_ip_7 = len(lpbck_ip_7)
lenlpbck_ip_8 = len(lpbck_ip_8)
lenlpbck_ip_9 = len(lpbck_ip_9)
lenlpbck_ip_0 = len(lpbck_ip_0)
lenlpbck_ip_A = len(lpbck_ip_A)
lenlpbck_ip_B = len(lpbck_ip_B)
lenlpbck_name = len(lpbck_name)
print("\n\n**** Number of list entries (should be all the same):")
print(lenlpbkint_id)
print(lenvrf___name)
print(lenlpbck_ip_1)
print(lenlpbck_ip_2)
print(lenlpbck_ip_3)
print(lenlpbck_ip_4)
print(lenlpbck_ip_5)
print(lenlpbck_ip_6)
print(lenlpbck_ip_7)
print(lenlpbck_ip_8)
print(lenlpbck_ip_9)
print(lenlpbck_ip_0)
print(lenlpbck_ip_A)
print(lenlpbck_ip_B)
print(lenlpbck_name)

# function definition
def switch_fnc(switch_idx, lpbck_ip_x):
	if not "NONE" in switch_idx:
		outfil = "04-out-vrf-loopback-ip-" + switch_idx + ".cfg"
		print("\nThe name of the out-file is " + outfil + " \n")
		answer()
		with open(outfil, 'w') as outfile:
			outfile.write('\n# ***** start of ' + outfil + ' ***** ')
		with open(outfil, 'a') as outfile:
			outfile.write('\n\n')
		script = "rwa"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("rwa\n")
		script = "rwa"
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("rwa\n")
		script = "enable " + " "
		print (script)
		with open(outfil, 'a') as outfile:
			outfile.write("enable " + "\n\n")
		script = "configure terminal " + " "
		print (script)
		with open(outfil, 'a') as outfile:
			outfile.write("configure terminal\n\n")
		
		for value in range(0,lenlpbkint_id):
			if lpbck_ip_x[value]:
				if int(lpbkint_id[value]) >= 0:
					script = "interface loopback " + lpbkint_id[value] + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("interface loopback " + lpbkint_id[value] + " \n")
					if "GRT" in vrf___name[value]:
						if ":" in lpbck_ip_x[value]:
							script = 'ipv6 interface address ' + lpbck_ip_x[value] + '/128'
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('ipv6 interface address ' + lpbck_ip_x[value] + '/128\n')
							script = 'ipv6 interface name "CLIPv6-' + vrf___name[value] + '"'
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('ipv6 interface name "CLIPv6-' + vrf___name[value] + '" \n')
						else:	
							script = 'ip address ' + lpbck_ip_x[value] + '/32' + ' name "CLIP-' + vrf___name[value] + '"'
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('ip address ' + lpbck_ip_x[value] + '/32' + ' name "CLIP-' + vrf___name[value] + '" \n')
					else:
						if ":" in lpbck_ip_x[value]:
							script = 'ipv6 interface address ' + lpbck_ip_x[value] + '/128 vrf ' + vrf___name[value]
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('ipv6 interface address ' + lpbck_ip_x[value] + '/128 vrf ' + vrf___name[value] + '\n')
							script = 'ipv6 interface name "CLIPv6-' + vrf___name[value] + '" vrf ' + vrf___name[value]
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('ipv6 interface name "CLIPv6-' + vrf___name[value] + '" vrf ' + vrf___name[value] + '\n')
						else:
							script = 'ip address ' + lpbck_ip_x[value] + '/32 vrf ' + vrf___name[value] + ' name "CLIP-' + vrf___name[value] + '"'
							print(script)
							with open(outfil, 'a') as outfile:
								outfile.write('ip address ' + lpbck_ip_x[value] + '/32 vrf ' + vrf___name[value] + ' name "CLIP-' + vrf___name[value] + '" \n')
					script = "exit " + " "
					print(script)
					with open(outfil, 'a') as outfile:
						outfile.write("exit " + " \n\n")
		
		script = "\nend " + " "
		print(script)
		with open(outfil, 'a') as outfile:
			outfile.write("\nend " + "\n\n")
		script = "save config " + " "
		print (script)
		with open(outfil, 'a') as outfile:
			outfile.write("save config " + " \n\n")
		script = "backup configure 04-loopback-ip-vrf " + " "
		print (script)
		with open(outfil, 'a') as outfile:
			outfile.write("backup configure 04-loopback-ip-vrf " + " \n\n")
		with open(outfil, 'a') as outfile:
			outfile.write('\n# ***** end of ' + outfil + ' ***** \n\n')

# build dictionary

switch_lib = {}

switch_lib = {
	switch_id1 : lpbck_ip_1,
	switch_id2 : lpbck_ip_2,
	switch_id3 : lpbck_ip_3,
	switch_id4 : lpbck_ip_4,
	switch_id5 : lpbck_ip_5,
	switch_id6 : lpbck_ip_6,
	switch_id7 : lpbck_ip_7,
	switch_id8 : lpbck_ip_8,
	switch_id9 : lpbck_ip_9,
	switch_id0 : lpbck_ip_0,
	switch_idA : lpbck_ip_A,
	switch_idB : lpbck_ip_B,
	}

script = "\n ***** The complete library looks like: " + "\n"
print(script)
print(switch_lib)
script = "\n ***** Building configuration files: " + "\n"
print(script)
answer()

for key, value in switch_lib.items():
	print("\nSwitch: " + key + "\tLoopback IP address(es): ")
	for values in value:
		print("\t" + values)
	switch_fnc(key, value)

script = "\n ***** program end ***** " + " "
print(script)

